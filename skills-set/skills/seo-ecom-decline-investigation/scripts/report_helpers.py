#!/usr/bin/env python
"""
report_helpers.py — reusable branded-openpyxl primitives for the Phase 8 XLSX build.

Import this module rather than re-deriving these functions — it has two real bugs from the
source investigation's report build already found and fixed, both documented in
references/pitfalls.md #12 and #13. Re-deriving these primitives from scratch risks
re-introducing the same bugs.

Colors follow built-in report branding palette exactly — see that skill for the full spec
(logo processing, DOCX/HTML patterns, etc.). This module only covers the XLSX primitives.

No external AI/LLM API is called anywhere in this file.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

YELLOW = "F5C518"
BLACK = "0A0A0A"
DARK = "1A1A1A"
WHITE = "FFFFFF"
MUTED = "888888"
GREEN = "2ECC71"
RED = "E74C3C"
LIGHT = "F2F2F2"
TEXT = "1C1C1C"
ORANGE = "E67E22"


def thin(color="DDDDDD"):
    return Border(*[Side(style="thin", color=color)] * 4)


def all_border(style="thin", color="333333"):
    return Border(*[Side(style=style, color=color)] * 4)


def header_band(ws, title, meta, logo_path, last_col=12):
    """Black header band, rows 1-5, with a yellow divider on row 6 and the processed logo at B2.
    logo_path is optional; pass None/empty for text-only branding (no logo)."""
    for r in range(1, 6):
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BLACK)
        ws.row_dimensions[r].height = 14
    for c in range(1, last_col + 1):
        ws.cell(row=6, column=c).fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[6].height = 4
    if logo_path:
        logo = XLImage(logo_path)
        logo.width = 161
        logo.height = 60
        ws.add_image(logo, "B2")
    t = ws.cell(row=2, column=5, value=title)
    t.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    t.fill = PatternFill("solid", fgColor=BLACK)
    t.alignment = Alignment(horizontal="left", vertical="center")
    m = ws.cell(row=4, column=5, value=meta)
    m.font = Font(name="Arial", size=9, color=MUTED)
    m.fill = PatternFill("solid", fgColor=BLACK)
    m.alignment = Alignment(horizontal="left", vertical="center")


def section(ws, row, text, last_col=12):
    """A yellow-left-border section heading row. Returns the next free row."""
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=2, value=text)
    c.font = Font(name="Arial", bold=True, size=13, color=TEXT)
    c.border = Border(left=Side(style="thick", color=YELLOW), bottom=Side(style="thin", color="EEEEEE"))
    ws.row_dimensions[row].height = 26
    return row + 2


def subnote(ws, row, text, last_col=12, italic=True, size=9, color=MUTED):
    """A muted italic footnote row (e.g. methodology caveats). Returns the next free row."""
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=2, value=text)
    c.font = Font(name="Arial", italic=italic, size=size, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def header_row(ws, row, col, values, widths=None):
    """Black-background, yellow-text table header row.

    IMPORTANT (pitfalls.md #14): if a sheet has multiple tables with DIFFERENT column counts,
    do not pass `widths` per-table — column widths are a worksheet-level property and later
    calls silently overwrite earlier ones. Set widths ONCE at the end of the sheet, sized to a
    layout that works for every table on that sheet (pad narrower tables with empty header
    cells rather than fighting the column count)."""
    for i, v in enumerate(values):
        cc = ws.cell(row=row, column=col + i, value=v)
        cc.font = Font(name="Arial", bold=True, size=10, color=YELLOW)
        cc.fill = PatternFill("solid", fgColor=BLACK)
        cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cc.border = all_border()
        if widths:
            ws.column_dimensions[get_column_letter(col + i)].width = widths[i]
    ws.row_dimensions[row].height = 30
    return row + 1


def data_row(ws, row, col, values, accent=None, bold_first=False, shade=None):
    """Alternating-shade data row with an optional colored left-border accent (severity/status).

    IMPORTANT (pitfalls.md #13): `accent` must be one of the color VARIABLES above (GREEN, RED,
    ORANGE, YELLOW...), never a bare string like "GREEN". Passing a string crashes deep inside
    openpyxl's Border/Side color setter with a confusing "Colors must be aRGB hex values" error
    far from the actual typo. If authoring a large table by hand, grep the source for stray
    quoted color-name strings before running it."""
    sh = shade if shade else (WHITE if row % 2 == 0 else LIGHT)
    for i, v in enumerate(values):
        cc = ws.cell(row=row, column=col + i, value=v)
        cc.font = Font(name="Arial", bold=(bold_first and i == 0), size=10, color=TEXT)
        cc.fill = PatternFill("solid", fgColor=sh)
        cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if accent and i == 0:
            cc.border = Border(left=Side(style="medium", color=accent), right=Side(style="thin", color="DDDDDD"),
                                top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))
        else:
            cc.border = thin()
    ws.row_dimensions[row].height = 34
    return row + 1


def stat_card(ws, row, col, number, change, label, change_color=GREEN):
    """A 2-column-wide dark stat card: big yellow number, colored change indicator, muted label.

    IMPORTANT (pitfalls.md #12): merge each LINE separately (number / change / label as three
    distinct 1-row x 2-col merges), never merge the whole multi-row card as one block. openpyxl
    only exposes the top-left cell of a merged range as writable — merging rows [row, row+3] as
    one cell and then trying to write to row+2 and row+3 individually throws
    `AttributeError: 'MergedCell' object attribute 'value' is read-only`."""
    for cc in [col, col + 1]:
        ws.cell(row=row - 1, column=cc).fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[row - 1].height = 5
    for rr in range(row, row + 4):
        for cc in [col, col + 1]:
            ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=DARK)

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    n = ws.cell(row=row, column=col, value=number)
    n.font = Font(name="Arial", bold=True, size=18, color=YELLOW)
    n.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)
    ch = ws.cell(row=row + 2, column=col, value=change)
    ch.font = Font(name="Arial", bold=True, size=9, color=change_color)
    ch.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=row + 3, start_column=col, end_row=row + 3, end_column=col + 1)
    lb = ws.cell(row=row + 3, column=col, value=label.upper())
    lb.font = Font(name="Arial", size=8, color=MUTED)
    lb.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def finalize(ws, freeze="B8"):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.freeze_panes = freeze


def process_logo(source_url, out_path, pad=200):
    """Optional legacy helper — unused when shipping without a logo."""
    import io
    import requests
    import numpy as np
    from PIL import Image

    resp = requests.get(source_url, timeout=15)
    img = Image.open(io.BytesIO(resp.content)).convert("RGBA")
    arr = np.array(img)
    r, g, b = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2]
    non_black = (r > 20) | (g > 20) | (b > 20)
    rows = np.where(np.any(non_black, axis=1))[0]
    cols = np.where(np.any(non_black, axis=0))[0]
    cropped = img.crop((
        max(0, cols[0] - pad), max(0, rows[0] - pad),
        min(img.size[0], cols[-1] + pad), min(img.size[1], rows[-1] + pad),
    ))
    cropped.save(out_path)
    return out_path

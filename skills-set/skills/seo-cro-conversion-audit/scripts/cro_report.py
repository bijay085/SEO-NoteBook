#!/usr/bin/env python3
"""Assemble the CRO audit workbook from the analysis JSONs.

Usage:
    python cro_report.py <analysis_dir> <output.xlsx>

Reads whatever exists in <analysis_dir>: cro_signals.json, form_audit.json,
clarity_findings.json, findings.json, cro_verdict.json. Every sheet is gated on
its input being present, so a partial run still produces a valid workbook.

Styling here is clean and neutral. Final client-facing brand polish (colors,
logo, the narrative HTML/DOCX) is applied by the built-in report branding
skill — this script does not hard-code a brand palette it cannot verify.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common import read_json

# Neutral placeholder palette — the branding skill overrides for delivery.
BRAND_PRIMARY = "1F3A5F"
BRAND_LIGHT = "EAF0F6"
WHITE = "FFFFFF"

_HEAD = Font(bold=True, color=WHITE, size=11)
_TITLE = Font(bold=True, color=BRAND_PRIMARY, size=16)
_BOLD = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_FILL = PatternFill("solid", fgColor=BRAND_PRIMARY)
_LIGHT = PatternFill("solid", fgColor=BRAND_LIGHT)
_thin = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _load(d, name):
    p = Path(d) / name
    return read_json(p) if p.exists() else None


def _final(host, draft, verdict):
    """Corrected verdict scores override the mechanical draft when present."""
    if verdict and isinstance(verdict.get("scores"), dict) and verdict["scores"].get(host):
        return verdict["scores"][host]
    return draft


def _headers(ws, row, labels, widths=None):
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font, cell.fill, cell.border = _HEAD, _FILL, _BORDER
        cell.alignment = _WRAP
        if widths:
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]


def _row(ws, row, values, wrap=False):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = _BORDER
        if wrap:
            cell.alignment = _WRAP


def _overview(wb, sig, verdict, findings):
    ws = wb.active
    ws.title = "Overview"
    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16
    ws.cell(row=1, column=1, value="CRO / Conversion Audit").font = _TITLE
    r = 3
    site = (sig or {}).get("site_domain", "")
    ws.cell(row=r, column=1, value="Site").font = _BOLD
    ws.cell(row=r, column=2, value=site); r += 1
    if verdict and verdict.get("headline"):
        ws.cell(row=r, column=1, value="Headline").font = _BOLD
        c = ws.cell(row=r, column=2, value=verdict["headline"]); c.alignment = _WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 90
        r += 2
    if sig:
        _headers(ws, r, ["Domain"] + sig["dimensions"] + ["Overall"]); r += 1
        for host, d in sig["domains"].items():
            sc = _final(host, d["scores"], verdict)
            _row(ws, r, [host] + [sc.get(dim) for dim in sig["dimensions"]] + [sc.get("Overall")])
            if d.get("is_site"):
                for c in range(1, len(sig["dimensions"]) + 2):
                    ws.cell(row=r, column=c).fill = _LIGHT
            r += 1
        r += 1
    if findings:
        ws.cell(row=r, column=1, value="Top findings").font = _BOLD; r += 1
        _headers(ws, r, ["Severity", "Area", "Finding"], [12, 12, 90]); r += 1
        for f in findings[:12]:
            _row(ws, r, [f.get("severity"), f.get("area"), f.get("title")], wrap=True); r += 1


def _scores(wb, sig, verdict):
    if not sig:
        return
    ws = wb.create_sheet("Conversion Scores")
    _headers(ws, 1, ["Domain"] + sig["dimensions"] + ["Overall"],
             [30] + [12] * (len(sig["dimensions"]) + 1))
    r = 2
    for host, d in sig["domains"].items():
        sc = _final(host, d["scores"], verdict)
        _row(ws, r, [host] + [sc.get(dim) for dim in sig["dimensions"]] + [sc.get("Overall")])
        r += 1


def _matrix(wb, sig):
    if not sig:
        return
    ws = wb.create_sheet("Comparison Matrix")
    hosts = list(sig["domains"].keys())
    _headers(ws, 1, ["Signal"] + hosts, [26] + [22] * len(hosts))
    keys = ["pages", "service_pages", "has_form", "has_sticky", "phone_present",
            "hero_frac", "avg_words", "pricing", "testimonials", "guarantee",
            "badges", "review_counts", "specific_cta", "top_ctas"]
    r = 2
    for k in keys:
        vals = []
        for h in hosts:
            v = sig["domains"][h]["profile"].get(k)
            vals.append(", ".join(map(str, v)) if isinstance(v, list) else v)
        _row(ws, r, [k] + vals, wrap=True); r += 1


def _form(wb, fa):
    if not fa or not fa.get("found"):
        return
    ws = wb.create_sheet("Form Audit")
    _headers(ws, 1, ["Field", "Tag", "Type", "Role", "Required", "Maxlength"],
             [30, 10, 10, 12, 12, 12])
    r = 2
    for fld in fa["fields"]:
        _row(ws, r, [fld["label"], fld["tag"], fld["type"], fld["role"],
                     "yes" if fld["required"] else "no", fld["maxlength"]]); r += 1
    ws2 = wb.create_sheet("Form Test Plan")
    _headers(ws2, 1, ["ID", "Category", "Scenario", "Target", "Expected",
                      "Actual", "Verdict", "Fix"], [8, 16, 34, 10, 40, 20, 12, 24])
    r = 2
    for tc in fa["test_plan"]:
        _row(ws2, r, [tc["id"], tc["category"], tc["scenario"], tc["target_field"],
                      tc["expected"], tc["actual"], tc["verdict"], tc["fix"]], wrap=True)
        r += 1


def _behavioral(wb, clarity):
    if not clarity or not clarity.get("pages"):
        return
    ws = wb.create_sheet("Behavioral (Clarity)")
    _headers(ws, 1, ["Page", "Page views", "Conversion-click %", "Biggest scroll drop",
                     "Top distractor (clicks)"], [24, 12, 18, 30, 40])
    r = 2
    for label, rec in clarity["pages"].items():
        click = rec.get("click", {})
        scroll = rec.get("scroll", {})
        drop = scroll.get("biggest_single_band_drop") or {}
        drop_s = (f"{drop.get('from_reach')}%->{drop.get('to_reach')}% "
                  f"({drop.get('from_depth')}-{drop.get('to_depth')}% depth)" if drop else "")
        dist = click.get("top_distractor") or {}
        dist_s = f"{dist.get('element','')} ({dist.get('clicks','')})" if dist else ""
        _row(ws, r, [label, click.get("page_views"), click.get("conversion_click_pct"),
                     drop_s, dist_s], wrap=True); r += 1
    if clarity.get("integrity_issues"):
        r += 1
        ws.cell(row=r, column=1, value="Data-integrity flags").font = _BOLD; r += 1
        for it in clarity["integrity_issues"]:
            _row(ws, r, [it.get("issue"), it.get("detail")], wrap=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5); r += 1


def _actions(wb, verdict, findings):
    """Action Plan: every row carries Finding, Evidence, Solution, Executable steps."""
    recs = (verdict or {}).get("recommendations")
    ws = wb.create_sheet("Action Plan")
    _headers(ws, 1, ["#", "Priority", "Area", "Finding", "Evidence", "Impact",
                     "Solution", "Executable steps"], [4, 10, 10, 34, 36, 26, 34, 46])
    r = 2
    if recs:
        for i, rec in enumerate(recs, 1):
            steps = rec.get("steps") or []
            steps_s = "\n".join(f"{j}. {s}" for j, s in enumerate(steps, 1))
            solution = rec.get("solution") or rec.get("action") or ""
            finding = rec.get("finding") or ""
            _row(ws, r, [i, rec.get("priority"), rec.get("area"), finding,
                         rec.get("evidence"), rec.get("impact", ""), solution, steps_s], wrap=True)
            ws.row_dimensions[r].height = max(70, 15 * max(len(steps), 3))
            r += 1
    elif findings:
        for i, f in enumerate(findings, 1):
            acts = f.get("recommended_actions") or []
            _row(ws, r, [i, f.get("severity"), f.get("area"), f.get("title"),
                         f.get("evidence"), "", acts[0] if acts else "", ""], wrap=True)
            r += 1


def build(analysis_dir, out_xlsx):
    sig = _load(analysis_dir, "cro_signals.json")
    fa = _load(analysis_dir, "form_audit.json")
    clarity = _load(analysis_dir, "clarity_findings.json")
    findings = _load(analysis_dir, "findings.json") or []
    verdict = _load(analysis_dir, "cro_verdict.json")
    wb = Workbook()
    _overview(wb, sig, verdict, findings)
    _scores(wb, sig, verdict)
    _matrix(wb, sig)
    _form(wb, fa)
    _behavioral(wb, clarity)
    _actions(wb, verdict, findings)
    Path(out_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return [ws.title for ws in wb.worksheets]


def main(argv):
    if len(argv) < 2:
        print(__doc__); return 2
    sheets = build(argv[0], argv[1])
    print(f"cro_report: wrote {argv[1]} with sheets {sheets}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

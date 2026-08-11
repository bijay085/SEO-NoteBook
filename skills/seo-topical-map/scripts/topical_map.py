#!/usr/bin/env python3
"""Topical Map (Claude-native) : source extraction + deterministic branded export.

The ONLY Python in this skill. Claude does all ontology + QDP reasoning in-context and
writes an `ontology.json`; this script just (a) pulls clean text from DOCX/URL sources
and (b) renders the branded deliverables. No LLM/scraper deps.

    python topical_map.py sources --docx a.docx --url https://x --out corpus.txt
    python topical_map.py export --in ontology.json --out ./Deliverables

Deps: openpyxl (export only). Everything else is stdlib.
"""
import sys, os, json, re, argparse, zipfile, html, urllib.request
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from charts import chart_html
except Exception:
    def chart_html(*a, **k) -> str:
        return ""

YELLOW="F5C518"; BLACK="0A0A0A"; DARK="1A1A1A"; WHITE="FFFFFF"; MUT="888888"
LIGHT="F2F2F2"; TEXT="1C1C1C"; GREEN="2ECC71"
UA=("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36")

# ---------------------------------------------------------------- sources
def docx_text(path):
    with zipfile.ZipFile(path) as z:
        xml=z.read("word/document.xml").decode("utf-8","replace")
    xml=re.sub(r"</w:p>", "\n", xml)
    xml=re.sub(r"<[^>]+>", "", xml)
    return html.unescape(xml)

def url_text(url):
    req=urllib.request.Request(url, headers={"User-Agent":UA})
    with urllib.request.urlopen(req, timeout=30) as r:
        raw=r.read().decode("utf-8","replace")
    raw=re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", raw, flags=re.S|re.I)
    raw=re.sub(r"<[^>]+>", " ", raw)
    return html.unescape(re.sub(r"[ \t]+", " ", raw))

def plain_text(path):
    with open(path, encoding="utf-8", errors="replace") as f:
        t=f.read()
    if path.lower().endswith((".html",".htm")):
        t=re.sub(r"<[^>]+>"," ",t); t=html.unescape(t)
    return t

def cmd_sources(a):
    parts=[]
    for d in a.docx or []:
        parts.append(f"\n===== DOCX: {os.path.basename(d)} =====\n"+docx_text(d)); print("read docx", d)
    for u in a.url or []:
        try: parts.append(f"\n===== URL: {u} =====\n"+url_text(u)); print("fetched", u)
        except Exception as e: print(" URL failed", u, e)
    for t in a.txt or []:
        parts.append(f"\n===== FILE: {os.path.basename(t)} =====\n"+plain_text(t)); print("read", t)
    text="\n".join(parts).strip()
    if a.out:
        open(a.out,"w",encoding="utf-8").write(text); print(f"\nSAVED {a.out} ({len(text):,} chars)")
    else:
        print(text[:4000])

# ---------------------------------------------------------------- export
FULL=["#","Relevance Layer","Entity Name","Type","Salience","Ontology Layer",
      "Relationship Type","Short Definition","Entity Definition","Relational Definition"]
ATTR=["Entity","Attribute","Business Attribute","Value","Entity Role","Buyer Context",
      "Commercial Relevance","Template Assigned (Value)","Template","Template Family",
      "Section","Keyword","SV","KD","CPC","Slug","Title","KW Source","Page Status",
      "KW Fit","Priority","KW Rejection Note"]
ATTR_KEYS=["entity","attribute","business_attribute","value","entity_role","buyer_context",
      "commercial_relevance","template_assigned","template","template_family","section",
      "keyword","sv","kd","cpc","slug","title","kw_source","page_status","kw_fit",
      "priority","kw_rejection_note"]
KW=["#","Entity Name","Kind","Relevance Layer","Salience","Primary KW?","Keyword",
    "Volume","KD","CPC","Intent","Trend","KW Type"]
# --- page plan (QDP decision) + review (noise) ---
PAGES_COLS=["#","Entity","Attribute","Value","Decision","Merge Into","Section","Page Type",
    "Title","Slug","Primary Keyword","SV","KD","QDP","Rules (V·E·S·P)","SERP","Note"]
BLOG_COUNT_COLS=["Entity","Attribute","Value","Blog Count","Landing Count","Total","Noise Removed"]
REVIEW_COLS=["Entity","Attribute","Value","Category","Title","Slug","Keyword","SV","Page Type","Reason"]

def _qdp_rules(q):
    if not q: return ""
    def m(v): return "?" if v is None else ("✓" if v else "✗")
    return (f"V{m(q.get('volume'))} E{m(q.get('different_entity'))} "
            f"S{m(q.get('low_similarity'))} P{m(q.get('pattern'))}")

def _page_rows(pages):
    out=[]
    for i,p in enumerate(pages,1):
        q=p.get("qdp",{}) or {}
        out.append([i, p.get("entity",""), p.get("attribute",""), p.get("value",""),
            p.get("decision",""), p.get("merge_into","") or "", p.get("section",""),
            p.get("page_type",""), p.get("title",""), p.get("slug",""),
            p.get("primary_keyword","") or p.get("keyword",""), p.get("sv",""), p.get("kd",""),
            (f"{q.get('score','')}/4" if q else ""), _qdp_rules(q),
            q.get("serp_verdict",""), p.get("note","")])
    return out

def _blog_count_rows(pages, review):
    agg={}; noise={}
    for p in pages:
        if p.get("decision")=="page":
            k=(p.get("entity",""),p.get("attribute",""),p.get("value",""))
            a=agg.setdefault(k,{"blog":0,"landing":0})
            if str(p.get("page_type","")).lower().startswith("blog"): a["blog"]+=1
            else: a["landing"]+=1
    for r in review:
        k=(r.get("entity",""),r.get("attribute",""),r.get("value",""))
        noise[k]=noise.get(k,0)+1
    rows=[]
    for k in sorted(set(list(agg)+list(noise))):
        a=agg.get(k,{"blog":0,"landing":0}); n=noise.get(k,0)
        rows.append([k[0],k[1],k[2],a["blog"],a["landing"],a["blog"]+a["landing"],n])
    return rows

def _review_rows(review):
    return [[r.get("entity",""),r.get("attribute",""),r.get("value",""),r.get("category",""),
        r.get("title",""),r.get("slug",""),r.get("keyword",""),r.get("sv",""),
        r.get("page_type",""),r.get("reason","")] for r in review]

def _sheet(wb, title, headers, rows, tab=YELLOW):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    ws=wb.create_sheet(title); ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=tab
    for i,h in enumerate(headers,1):
        c=ws.cell(1,i,h); c.font=Font(name="Arial",bold=True,size=10,color=YELLOW)
        c.fill=PatternFill("solid",fgColor=BLACK); c.alignment=Alignment(vertical="center",wrap_text=True)
        _side=Side(style="thin",color="333333")
        c.border=Border(left=_side, right=_side, top=_side, bottom=_side)
        ws.column_dimensions[get_column_letter(i)].width=min(46,max(10,len(h)+4))
    ws.row_dimensions[1].height=28
    for r,row in enumerate(rows,2):
        sh=WHITE if r%2 else LIGHT
        for i,v in enumerate(row,1):
            c=ws.cell(r,i,v); c.font=Font(name="Arial",size=10,color=TEXT)
            c.fill=PatternFill("solid",fgColor=sh); c.alignment=Alignment(vertical="top",wrap_text=True)
            _side=Side(style="thin",color="DDDDDD")
            c.border=Border(left=_side, right=_side, top=_side, bottom=_side)
    ws.freeze_panes="A2"; return ws

def cmd_export(a):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    d=json.load(open(a.inp, encoding="utf-8"))
    ce=d.get("central_entity",{}); ents=d.get("entities",[]); attrs=d.get("attributes",[]); kws=d.get("keywords",[])
    pages=d.get("pages",[]); review=d.get("review",[])
    os.makedirs(a.out, exist_ok=True)

    wb=Workbook()
    _ws=wb.active
    if _ws is not None:
        wb.remove(_ws)
    # Topical Map (overview)
    ov=wb.create_sheet("Topical Map"); ov.sheet_view.showGridLines=False; ov.sheet_properties.tabColor=YELLOW
    ov.column_dimensions['A'].width=26; ov.column_dimensions['B'].width=70
    def kv(r,k,v):
        a1=ov.cell(r,1,k); a1.font=Font(name="Arial",bold=True,size=10,color=TEXT)
        b1=ov.cell(r,2,v); b1.font=Font(name="Arial",size=10,color=TEXT); b1.alignment=Alignment(wrap_text=True,vertical="top")
    t=ov.cell(1,1,"Topical Authority Map"); t.font=Font(name="Arial",bold=True,size=15,color=TEXT)
    ov.cell(2,1,"SEO").font=Font(name="Arial",bold=True,size=10,color="B8890A")
    kv(4,"Central Entity",ce.get("name","")); kv(5,"Type",ce.get("type","")); kv(6,"Domain",ce.get("domain",""))
    kv(7,"Definition",ce.get("definition","")); kv(8,"Summary",d.get("summary",""))
    kv(10,"Entities",str(len(ents))); kv(11,"Attribute rows",str(len(attrs))); kv(12,"Keyword rows",str(len(kws)))
    layers={}
    for e in ents: layers[e.get("relevance_layer","?")]=layers.get(e.get("relevance_layer","?"),0)+1
    kv(13,"By relevance layer", ", ".join(f"{k}: {v}" for k,v in sorted(layers.items())))
    BT={1:"Types",2:"Components/Parts",3:"Methods",4:"Services",5:"Target Problems",6:"Tools/Products",
        7:"People/Roles",8:"Providers",9:"Regulators",10:"Certifications",11:"By Location",12:"Concepts",13:"Domain-Specific"}
    bk={}
    for e in ents:
        b=e.get("primary_bucket")
        if b: bk[b]=bk.get(b,0)+1
    kv(14,"By bucket (primary)", ", ".join(f"{BT.get(k,k)}: {v}" for k,v in sorted(bk.items())))
    # page-plan summary (only when a page plan exists)
    if pages:
        dec={}
        for p in pages: dec[p.get("decision","?")]=dec.get(p.get("decision","?"),0)+1
        kv(15,"Page decisions", ", ".join(f"{k}: {v}" for k,v in sorted(dec.items())))
        bl={"blog":0,"landing":0}
        for p in pages:
            if p.get("decision")=="page":
                if str(p.get("page_type","")).lower().startswith("blog"): bl["blog"]+=1
                else: bl["landing"]+=1
        kv(16,"Pages (Blog / Landing)", f"{bl['blog']} / {bl['landing']}")
    if review: kv(17,"Noise removed", str(len(review)))

    _sheet(wb,"Full Ontology",FULL,[[i, e.get("relevance_layer",""), e.get("entity_name",""),
        e.get("entity_type",""), e.get("salience",""), e.get("ontology_layer",""),
        e.get("relationship_type",""), e.get("short_definition",""), e.get("entity_definition",""),
        e.get("relational_definition","")] for i,e in enumerate(ents,1)], tab=GREEN)
    _sheet(wb,"Attribute Map",ATTR,[[r.get(k,"") for k in ATTR_KEYS] for r in attrs], tab="3498DB")
    _sheet(wb,"Keyword Data",KW,[[i, k.get("entity_name",""), k.get("kind",""), k.get("relevance_layer",""),
        k.get("salience",""), ("Yes" if k.get("primary_kw") else ""), k.get("keyword",""), k.get("volume",""),
        k.get("kd",""), k.get("cpc",""), k.get("intent",""), k.get("trend",""), k.get("kw_type","")]
        for i,k in enumerate(kws,1)], tab="E67E22")
    # merged page-planning sheets (additive; skipped if absent)
    if pages: _sheet(wb,"Page Plan (AMR+QDP)",PAGES_COLS,_page_rows(pages), tab="9B59B6")
    if pages or review: _sheet(wb,"Blog Count by Value",BLOG_COUNT_COLS,_blog_count_rows(pages,review), tab="1ABC9C")
    if review: _sheet(wb,"Review (Noise Removed)",REVIEW_COLS,_review_rows(review), tab="E74C3C")

    xlsx=os.path.join(a.out,"entity_ontology.xlsx"); wb.save(xlsx)
    jpath=os.path.join(a.out,"entity_ontology.json"); json.dump(d, open(jpath,"w",encoding="utf-8"), indent=2, ensure_ascii=False)
    hpath=os.path.join(a.out,"entity_ontology.html"); open(hpath,"w",encoding="utf-8").write(_html(d,ce,ents,attrs,kws,BT))
    print("SAVED:", xlsx); print("SAVED:", jpath); print("SAVED:", hpath)
    if pages: print(f" page plan: {len(pages)} candidates")
    if review: print(f" review (noise): {len(review)} rows")

def _esc(x): return html.escape(str(x if x is not None else ""))
def _html(d,ce,ents,attrs,kws,BT):
    pages=d.get("pages",[])
    def _vol(x):
        try: return float(str(x).replace(",","").strip())
        except Exception: return None
    _kwp=[(k.get("keyword",""), _vol(k.get("volume"))) for k in kws]
    _kwp=[(l,v) for l,v in _kwp if v]; _kwp.sort(key=lambda t:-t[1])
    _kw_chart=chart_html("Top keywords by monthly search volume", _kwp[:12], "bars")
    def table(headers,rows):
        h="".join(f"<th>{_esc(x)}</th>" for x in headers)
        b="".join("<tr>"+"".join(f"<td>{_esc(c)}</td>" for c in r)+"</tr>" for r in rows)
        return f"<div class=tw><table><thead><tr>{h}</tr></thead><tbody>{b}</tbody></table></div>"
    ent_rows=[[i,e.get("relevance_layer",""),e.get("entity_name",""),e.get("entity_type",""),e.get("salience",""),
        BT.get(e.get("primary_bucket"),""),e.get("short_definition","")] for i,e in enumerate(ents,1)]
    kw_rows=[[k.get("entity_name",""),k.get("keyword",""),k.get("volume",""),k.get("kd",""),k.get("cpc",""),k.get("intent","")] for k in kws]
    page_section=""
    if pages:
        prows=[[p.get("decision",""),p.get("value",""),p.get("title",""),p.get("section",""),
            p.get("page_type",""),p.get("primary_keyword","") or p.get("keyword",""),p.get("sv",""),
            (f"{(p.get('qdp') or {}).get('score','')}/4")] for p in pages]
        page_section=("<h2>Page Plan (QDP)</h2>"
            +table(['Decision','Value','Title','Section','Type','Primary KW','SV','QDP'],prows))
    css=("body{font-family:Inter,Arial,sans-serif;background:#F7F7F7;color:#1C1C1C;margin:0}"
      ".h{background:#0A0A0A;border-bottom:3px solid #F5C518;padding:24px 40px}.h h1{color:#fff;margin:0;font-size:24px}"
      ".h h1 span{color:#F5C518}.h p{color:#888;margin:4px 0 0;font-size:13px}main{max-width:1100px;margin:0 auto;padding:0 32px 60px}"
      "h2{border-left:5px solid #F5C518;padding-left:12px;margin:36px 0 10px;font-size:18px}"
      ".card{background:#1A1A1A;color:#fff;border-radius:8px;padding:16px 20px;margin:18px 0}.card b{color:#F5C518}"
      ".tw{overflow-x:auto;border-radius:8px;box-shadow:0 2px 12px rgba(0,0,0,.08);margin:12px 0;background:#fff}"
      "table{width:100%;border-collapse:collapse;font-size:12.5px;min-width:640px}thead th{background:#0A0A0A;color:#F5C518;"
      "text-align:left;padding:9px 11px;font-size:10.5px;text-transform:uppercase}tbody td{padding:8px 11px;border-bottom:1px solid #eee;vertical-align:top}"
      "tbody tr:nth-child(even){background:#F6F6F6}")
    return (f"<!doctype html><html><head><meta charset=utf-8><meta name=viewport content='width=device-width,initial-scale=1'>"
      f"<title>Topical Map : {_esc(ce.get('name',''))} | SEO</title><style>{css}</style></head><body>"
      f"<div class=h><h1>Topical Authority <span>Map</span></h1><p>{_esc(ce.get('name',''))} &nbsp;|&nbsp; {_esc(ce.get('type',''))} &nbsp;|&nbsp; SEO</p></div><main>"
      f"<div class=card><b>Central entity:</b> {_esc(ce.get('name',''))} : {_esc(ce.get('definition',''))}<br><br>{_esc(d.get('summary',''))}"
      f"<br><br><b>{len(ents)}</b> entities · <b>{len(attrs)}</b> attribute rows · <b>{len(kws)}</b> keyword rows"
      + (f" · <b>{len(pages)}</b> page candidates" if pages else "") + "</div>"
      f"<h2>Entity Ontology</h2>{table(['#','Relevance Layer','Entity','Type','Salience','Bucket','Short Definition'],ent_rows)}"
      f"{page_section}"
      f"{_kw_chart}<h2>Keyword Data</h2>{table(['Entity','Keyword','Volume','KD','CPC','Intent'],kw_rows)}"
      f"</main></body></html>")

def main():
    p=argparse.ArgumentParser(); sub=p.add_subparsers(dest="cmd", required=True)
    s=sub.add_parser("sources"); s.add_argument("--docx",action="append"); s.add_argument("--url",action="append")
    s.add_argument("--txt",action="append"); s.add_argument("--out")
    e=sub.add_parser("export"); e.add_argument("--in",dest="inp",required=True); e.add_argument("--out",default="./Deliverables")
    a=p.parse_args()
    {"sources":cmd_sources,"export":cmd_export}[a.cmd](a)

if __name__=="__main__":
    main()

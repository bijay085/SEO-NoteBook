#!/usr/bin/env python3
"""report_helpers.py : branded openpyxl primitives for the master XLSX (SEO report branding,
config-driven). Mirrors the look of brand_lib.py's HTML so the two deliverables match.

    import report_helpers as R
    wb = R.new_book()
    ws = R.sheet(wb, "Overview")
    R.band(ws, "Client · Sandbox-Effect Analysis", sub="Prepared by Bijay")
    R.stat_cards(ws, 3, [("Graduation Score","48/100"),("Brand click share","97%"),("Non-brand pos","38")])
    R.section(ws, 7, "Suppression signature")
    R.table(ws, 8, ["Month","Clicks","Impr","CTR%"], rows)
    wb.save("out.xlsx")

Fonts default to Lexend (the client's stated preference across all deliverables); override via
config brand.font. Colors from config brand block. openpyxl only."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT={'yellow':'F5C518','black':'0A0A0A','dark':'1A1A1A','text':'1C1C1C','muted':'888888',
         'green':'2ECC71','red':'E74C3C','orange':'E67E22','font':'Lexend','agency':'Bijay'}
def _hex(x): return (x or '').lstrip('#').upper() or '000000'
def palette(cfg):
    b=dict(DEFAULT); b.update({k:_hex(v) if k not in ('font','agency') else v
                               for k,v in (cfg.get('brand',{}) if cfg else {}).items() if v})
    return b
_thin=Side(style='thin', color='E6E6E6')
BORDER=Border(left=_thin,right=_thin,top=_thin,bottom=_thin)

def new_book():
    wb=Workbook()
    ws=wb.active
    if ws is not None:
        wb.remove(ws)
    return wb
def sheet(wb, title): return wb.create_sheet(title[:31])

def band(ws, title, sub='', cfg=None, width=8):
    b=palette(cfg)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=width)
    c=ws.cell(1,1,title); c.fill=PatternFill('solid',fgColor=b['black'])
    c.font=Font(name=b['font'],size=15,bold=True,color=b['yellow']); c.alignment=Alignment('left','center',indent=1)
    ws.row_dimensions[1].height=30
    if sub:
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=width)
        s=ws.cell(2,1,sub); s.font=Font(name=b['font'],size=9,color=b['muted']); s.alignment=Alignment('left',indent=1)

def section(ws, row, title, cfg=None, width=8):
    b=palette(cfg)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=width)
    c=ws.cell(row,1,title); c.fill=PatternFill('solid',fgColor=b['dark'])
    c.font=Font(name=b['font'],size=12,bold=True,color='FFFFFF'); c.alignment=Alignment('left',indent=1)
    ws.row_dimensions[row].height=22

def stat_cards(ws, row, cards, cfg=None):
    b=palette(cfg)
    col=1
    for label,val in cards:
        ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+1)
        v=ws.cell(row,col,val); v.fill=PatternFill('solid',fgColor=b['dark'])
        v.font=Font(name=b['font'],size=16,bold=True,color=b['yellow']); v.alignment=Alignment('center','center')
        ws.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+1)
        l=ws.cell(row+1,col,label); l.fill=PatternFill('solid',fgColor=b['dark'])
        l.font=Font(name=b['font'],size=9,color='FFFFFF'); l.alignment=Alignment('center','center',wrap_text=True)
        col+=2
    ws.row_dimensions[row].height=26

def table(ws, row, headers, rows, cfg=None, widths=None):
    b=palette(cfg)
    for j,h in enumerate(headers,1):
        c=ws.cell(row,j,h); c.fill=PatternFill('solid',fgColor=b['black'])
        c.font=Font(name=b['font'],size=10,bold=True,color=b['yellow']); c.alignment=Alignment('left','center',wrap_text=True)
        c.border=BORDER
    ws.row_dimensions[row].height=20
    for i,r in enumerate(rows,1):
        for j,val in enumerate(r,1):
            c=ws.cell(row+i,j,val); c.font=Font(name=b['font'],size=10,color=b['text'])
            c.alignment=Alignment('left','top',wrap_text=True); c.border=BORDER
            if i%2==0: c.fill=PatternFill('solid',fgColor='F2F2F2')
    if widths:
        for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    return row+len(rows)+1

def autofit(ws, maxw=60):
    for col in ws.columns:
        letter=None; m=8
        for c in col:
            if c.column_letter: letter=c.column_letter
            try:
                if c.value: m=max(m,min(maxw,len(str(c.value))+2))
            except: pass
        if letter: ws.column_dimensions[letter].width=m

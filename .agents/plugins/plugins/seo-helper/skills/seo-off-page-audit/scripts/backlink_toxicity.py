#!/usr/bin/env python3
"""
backlink_toxicity.py - deterministic merge engine for the seo-off-page-audit skill.

Unions the toxic-link signal across sources and writes a conservative, domain-level
disavow file:
  - DataForSEO backlinks spam scores (saved JSON from mcp__dataforseo__backlinks_*),
  - a client Ahrefs export (CSV/XLSX) -- toxic/spam column if present,
  - a client Semrush toxic export (CSV/XLSX),
  - the existing disavow.txt (so nothing is re-disavowed).

A domain is a disavow candidate only when >= min_sources_to_disavow independent sources
flag it (Methodology sec.2). Single-source flags are downgraded to "review". Output is
the MERGED disavow file (existing + net-new), plus JSON for the report.

Credentials: loads project `.env` / host environment variables with a stdlib parser (python-dotenv is NOT installed).
This script performs no network calls itself -- the DataForSEO pull happens via the MCP
and is saved to JSON beforehand; here we only merge. No secret is written to any output.

Usage: python backlink_toxicity.py config.json
"""
import csv
import json
import os
import sys


def load_env(path="project `.env` / host environment variables"):
    env = {}
    p = os.path.expanduser(path)
    if not os.path.isfile(p):
        return env
    with open(p, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def read_config(path):
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def norm_domain(s):
    """Lowercase registrable-ish host: strip scheme, path, query, www, port."""
    s = (s or "").strip().lower()
    if not s:
        return ""
    if "://" in s:
        s = s.split("://", 1)[1]
    s = s.split("/")[0].split("?")[0].split(":")[0]
    if s.startswith("www."):
        s = s[4:]
    return s


def read_table(path):
    """Read a CSV or XLSX into a list of dicts with lowercased header keys."""
    p = os.path.expanduser(path)
    if not os.path.isfile(p):
        return []
    rows = []
    if p.lower().endswith(".xlsx"):
        from openpyxl import load_workbook # present; imported only when needed
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None) or []
        headers = [str(h).strip().lower() if h is not None else "" for h in header]
        for r in it:
            rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    else:
        with open(p, newline="", encoding="utf-8-sig", errors="replace") as fh:
            for r in csv.DictReader(fh):
                rows.append({(k or "").strip().lower(): v for k, v in r.items()})
    return rows


def find_col(headers, *needles):
    for h in headers:
        for n in needles:
            if n in h:
                return h
    return None


def to_float(v):
    try:
        return float(str(v).replace("%", "").strip())
    except (TypeError, ValueError):
        return None


def flags_from_table(path, score_needles, threshold, presence_is_flag):
    """Return {domain: score_or_'listed'} for rows this source considers toxic."""
    rows = read_table(path)
    if not rows:
        return {}
    headers = list(rows[0].keys())
    dcol = find_col(headers, "domain", "referring", "source url", "source")
    scol = find_col(headers, *score_needles) if score_needles else None
    if not dcol:
        sys.stderr.write("[off-page] no domain column in %s -- skipped\n" % path)
        return {}
    flagged = {}
    for r in rows:
        dom = norm_domain(str(r.get(dcol, "") or ""))
        if not dom:
            continue
        if scol:
            val = to_float(r.get(scol))
            if val is not None and val >= threshold:
                flagged[dom] = val
        elif presence_is_flag:
            flagged[dom] = "listed"
    return flagged


def walk_dfs(node, out):
    """Collect (domain, spam_score) from any nested DataForSEO JSON shape."""
    if isinstance(node, dict):
        dom = node.get("domain") or node.get("target")
        score = None
        for k in ("backlink_spam_score", "backlinks_spam_score", "spam_score"):
            if isinstance(node.get(k), (int, float)):
                score = node[k]
                break
        if dom and score is not None:
            out.append((norm_domain(str(dom)), float(score)))
        for v in node.values():
            walk_dfs(v, out)
    elif isinstance(node, list):
        for v in node:
            walk_dfs(v, out)


def dfs_flags(path, threshold):
    p = os.path.expanduser(path)
    if not p or not os.path.isfile(p):
        return {}
    try:
        with open(p, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except (json.JSONDecodeError, OSError):
        return {}
    pairs = []
    walk_dfs(data, pairs)
    return {d: s for d, s in pairs if s >= threshold}


def parse_disavow(path):
    p = os.path.expanduser(path or "")
    doms = set()
    if not p or not os.path.isfile(p):
        return doms
    with open(p, "r", encoding="utf-8", errors="replace") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if line.lower().startswith("domain:"):
                doms.add(norm_domain(line.split(":", 1)[1]))
            elif "." in line:
                doms.add(norm_domain(line))
    return doms


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: python backlink_toxicity.py config.json")
    cfg = read_config(sys.argv[1])
    _ = load_env(cfg.get("env_file", "project `.env` / host environment variables")) # loaded for parity; no net calls here
    ins = cfg.get("inputs", {})
    th = cfg.get("taxonomy", {}).get("toxic_thresholds", {})
    min_sources = int(th.get("min_sources_to_disavow", 2))
    out_dir = os.path.expanduser(cfg.get("output_dir", "./Off-Page-Audit/"))
    os.makedirs(out_dir, exist_ok=True)

    dfs = dfs_flags(ins.get("dataforseo_referring_domains_json", ""),
                    float(th.get("dataforseo_spam_score", 60)))
    ah = flags_from_table(ins.get("ahrefs_export", ""), ("toxic", "spam"),
                          float(th.get("ahrefs_toxicity", 30)), presence_is_flag=False)
    se = flags_from_table(ins.get("semrush_toxic_export", ""), ("toxic",),
                          float(th.get("semrush_toxic_score", 60)), presence_is_flag=True)
    already = parse_disavow(ins.get("existing_disavow", ""))

    per_source = {"dataforseo": dfs, "ahrefs": ah, "semrush": se}
    domains = set(dfs) | set(ah) | set(se)
    records = []
    for d in sorted(domains):
        srcs = [name for name, flags in per_source.items() if d in flags]
        verdict = "disavow-candidate" if len(srcs) >= min_sources else "review"
        if d in already:
            verdict = "already-disavowed"
        records.append({
            "domain": d,
            "sources": srcs,
            "source_count": len(srcs),
            "scores": {name: per_source[name][d] for name in srcs},
            "verdict": verdict,
        })

    new_disavow = sorted(r["domain"] for r in records
                         if r["verdict"] == "disavow-candidate" and r["domain"] not in already)
    merged = sorted(already | set(new_disavow))

    with open(os.path.join(out_dir, "disavow.txt"), "w", encoding="utf-8") as fh:
        fh.write("# Disavow file generated by seo-off-page-audit (merged: existing + high-confidence net-new)\n")
        fh.write("# High-confidence = flagged by >= %d independent sources. Review before uploading.\n" % min_sources)
        for d in merged:
            fh.write("domain:%s\n" % d)

    summary = {
        "domains_evaluated": len(domains),
        "already_disavowed": len(already),
        "disavow_candidates_total": sum(1 for r in records if r["verdict"] == "disavow-candidate"),
        "net_new_disavow": len(new_disavow),
        "review_single_source": sum(1 for r in records if r["verdict"] == "review"),
        "merged_disavow_size": len(merged),
        "sources_present": {k: bool(v) for k, v in per_source.items()},
    }
    with open(os.path.join(out_dir, "toxic_domains.json"), "w", encoding="utf-8") as fh:
        json.dump(records, fh, indent=2, ensure_ascii=False)
    with open(os.path.join(out_dir, "offpage_summary.json"), "w", encoding="utf-8") as fh:
        json.dump(summary, fh, indent=2, ensure_ascii=False)

    print("[off-page] evaluated=%d net_new_disavow=%d review=%d merged=%d sources=%s"
          % (summary["domains_evaluated"], summary["net_new_disavow"],
             summary["review_single_source"], summary["merged_disavow_size"],
             {k: v for k, v in summary["sources_present"].items() if v}))
    print("[off-page] wrote disavow.txt / toxic_domains.json / offpage_summary.json to %s" % out_dir)


if __name__ == "__main__":
    main()

from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from models import URLAuditResult

# ── Palette ───────────────────────────────────────────────────────────────────
C_DARK = "0F1117"
C_SUBHDR = "1E2440"
C_WHITE = "FFFFFF"
C_ROW_A = "F8F9FF"
C_ROW_B = "EEF0FF"
C_PASS = "22C55E"; C_PASS_BG = "DCFCE7"; C_PASS_FG = "166534"
C_WARN = "F59E0B"; C_WARN_BG = "FEF9C3"; C_WARN_FG = "854D0E"
C_CRIT = "EF4444"; C_CRIT_BG = "FEE2E2"; C_CRIT_FG = "991B1B"
C_INFO = "4F6EF7"; C_INFO_BG = "EEF2FF"; C_INFO_FG = "3730A3"
C_TEXT = "1E2235"


def _fill(h): return PatternFill("solid", fgColor=h)
def _font(sz=10, bold=False, color=C_TEXT): return Font(name="Arial", size=sz, bold=bold, color=color)
def _hfont(sz=11): return Font(name="Arial", size=sz, bold=True, color=C_WHITE)
def _border():
    s = Side(style="thin", color="D0D4E8")
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="left", wrap=False): return Alignment(horizontal=h, vertical="center", wrap_text=wrap, indent=1)

def _status_style(status):
    m = {
        "critical": (C_CRIT_BG, C_CRIT_FG),
        "warning": (C_WARN_BG, C_WARN_FG),
        "pass": (C_PASS_BG, C_PASS_FG),
        "info": (C_INFO_BG, C_INFO_FG),
    }
    return m.get(status.lower(), ("FFFFFF", C_TEXT))

def _header_row(ws, row, labels, bg=C_DARK):
    for i, lbl in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=lbl)
        c.font = _hfont()
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[row].height = 28

def _widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _banner(ws, text, cell, merge_to, bg=C_DARK, sz=14):
    ws.merge_cells(f"{cell}:{merge_to}")
    c = ws[cell]
    c.value = text
    c.font = Font(name="Arial", size=sz, bold=True, color=C_WHITE)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[int(''.join(filter(str.isdigit, cell)))].height = 40


def build_xlsx(results, output_path: str):
    wb = Workbook()
    _sheet_summary(wb, results)
    _sheet_action_plan(wb, results)
    _sheet_signal_detail(wb, results)
    _sheet_render_gap(wb, results)
    _sheet_bot_matrix(wb, results)
    _sheet_js_gated(wb, results)
    _sheet_missing_alt(wb, results)
    wb.save(output_path)


# ── Signal value display formatter ───────────────────────────────────────────
def _fmt_signal_val(v, signal_id: str = "") -> str:
    """
    Format a signal raw_value or rendered_value for display in a cell.
    Branches on signal_id first to avoid key-collision between images and links.
    """
    if v is None:
        return ": "
    if isinstance(v, dict):
        parts = []
        # body_text : identified by char_count key
        if signal_id == "body_text" or "char_count" in v:
            parts.append(f"{v.get('char_count', 0):,} chars")
            if v.get("gap_pct"):
                parts.append(f"gap {v['gap_pct']}%")
            gated = v.get("js_gated_text_samples", [])
            if gated:
                parts.append(f"{len(gated)} JS-gated blocks")
            elif not gated and signal_id == "body_text":
                parts.append("0 JS-gated blocks")
        # images : use signal_id OR signal_type key
        elif signal_id == "images" or v.get("signal_type") == "images":
            parts.append(f"{v.get('count', 0)} imgs")
            js = v.get("js_only_images", [])
            ma = v.get("missing_alt", [])
            lazy = v.get("lazy_load_detected", [])
            n_js = v.get("js_only_images_total", len(js))
            n_ma = v.get("missing_alt_total", len(ma))
            if n_js: parts.append(f"{n_js} JS-only")
            if n_ma: parts.append(f"{n_ma} missing alt")
            if lazy: parts.append(f"{len(lazy)} lazy-load")
        # internal_links
        elif signal_id == "internal_links" or "js_only_hrefs" in v:
            parts.append(f"{v.get('count', 0)} links")
            js = v.get("js_only_hrefs", [])
            if js:
                parts.append(f"{len(js)} JS-only")
        # json_ld
        elif signal_id == "json_ld" or "types" in v:
            types = v.get("types") or []
            parts.append(", ".join(types) if types else "none")
        else:
            parts = [f"{k}: {str(val)[:40]}" for k, val in list(v.items())[:2]]
        return " | ".join(parts) if parts else str(v)[:80]
    if isinstance(v, list):
        if not v:
            return ": "
        return ", ".join(str(x)[:30] for x in v[:5]) + \
               (f" (+{len(v)-5} more)" if len(v) > 5 else "")
    return str(v)
def _sheet_summary(wb, results):
    ws = wb.active
    ws.title = "Audit Summary"
    ws.sheet_view.showGridLines = False
    date_str = datetime.now().strftime("%B %d, %Y")

    _banner(ws, f"RENDERED vs RAW AUDIT REPORT | {date_str} | {len(results)} URLs", "A1", "I1")

    _header_row(ws, 2, [
        "URL", "Google Score", "AI Bot Score", "Render Gap Score",
        "Critical", "Warnings", "Passes", "robots.txt", "llms.txt"
    ])

    for r, res in enumerate(results, 3):
        bg = C_ROW_A if r % 2 == 1 else C_ROW_B
        scores = res.scores
        crit = sum(1 for s in res.signals if s.severity == "critical")
        warn = sum(1 for s in res.signals if s.severity == "warning")
        pas = sum(1 for s in res.signals if s.severity == "pass")

        row_data = [
            res.url,
            scores.get("google_score", 0),
            scores.get("ai_bot_score", 0),
            scores.get("render_gap_score", 0),
            crit, warn, pas,
            "✓ OK" if res.bot_access.get("Googlebot") == "allow" else "⚠ Check",
            "✓ Present" if res.llms_txt_status == "present" else "✗ Missing",
        ]
        ws.row_dimensions[r].height = 22
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.alignment = _align("left" if c == 1 else "center")

            if c in (2, 3, 4) and isinstance(val, int):
                if val >= 75: cell.font = _font(bold=True, color=C_PASS)
                elif val >= 50: cell.font = _font(bold=True, color=C_WARN)
                else: cell.font = _font(bold=True, color=C_CRIT)
            elif c == 5 and val > 0:
                cell.font = _font(bold=True, color=C_CRIT)
            else:
                cell.font = _font(bold=(c == 1))

    _widths(ws, {"A": 45, "B": 14, "C": 14, "D": 16,
                 "E": 11, "F": 11, "G": 9, "H": 16, "I": 13})


# ── Sheet 2 : Action Plan ─────────────────────────────────────────────────────
def _sheet_action_plan(wb, results):
    ws = wb.create_sheet("Action Plan")
    ws.sheet_view.showGridLines = False

    _banner(ws, "ACTION PLAN : Sorted by Priority", "A1", "I1", bg=C_SUBHDR, sz=13)
    _header_row(ws, 2, [
        "Rank", "URL", "Signal", "Category", "Severity",
        "Effort", "Impact", "Specific Action", "Verify"
    ], bg=C_SUBHDR)

    all_solutions = []
    for res in results:
        for sol in res.solutions:
            all_solutions.append((res.url, sol))
    all_solutions.sort(key=lambda x: x[1].priority_rank)

    for r, (url, sol) in enumerate(all_solutions, 3):
        bg = C_ROW_A if r % 2 == 1 else C_ROW_B
        ws.row_dimensions[r].height = None # auto-height for wrapped text
        row_data = [
            sol.priority_rank, url, sol.signal_id, sol.category,
            sol.severity.upper(),
            sol.effort, sol.impact, sol.fix, sol.verify,
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _border()
            cell.alignment = _align("left", wrap=(c in (8, 9)))
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=(c in (8, 9)), indent=1)
            if c == 5:
                bg2, fg = _status_style(sol.severity)
                cell.fill = _fill(bg2)
                cell.font = Font(name="Arial", size=9, bold=True, color=fg)
            else:
                cell.fill = _fill(bg)
                cell.font = _font(sz=9)

    _widths(ws, {"A": 6, "B": 35, "C": 18, "D": 14, "E": 11,
                 "F": 9, "G": 9, "H": 70, "I": 50})


# ── Sheet 3 : Signal Detail ───────────────────────────────────────────────────
def _sheet_signal_detail(wb, results):
    ws = wb.create_sheet("Signal Detail")
    ws.sheet_view.showGridLines = False

    _banner(ws, "SIGNAL-BY-SIGNAL: RAW vs RENDERED", "A1", "K1", bg=C_SUBHDR, sz=13)
    _header_row(ws, 2, [
        "URL", "Category", "Signal", "Raw Value", "Rendered Value",
        "Gap", "Match", "Severity", "Why This Severity", "Diagnosis", "Fix"
    ], bg=C_SUBHDR)

    # Build solution map : also cover sub-signal IDs
    sol_map = {}
    for res in results:
        for sol in res.solutions:
            sol_map[(res.url, sol.signal_id)] = sol
            # also map base signal id (e.g. images_missing_alt → images)
            base = sol.signal_id.replace("_missing_alt", "").replace("_js_gated", "")
            if (res.url, base) not in sol_map:
                sol_map[(res.url, base)] = sol

    r = 3
    for res in results:
        for sig in res.signals:
            bg = C_ROW_A if r % 2 == 1 else C_ROW_B
            sol = sol_map.get((res.url, sig.signal_id))
            # Auto-height for wrapped diagnosis/fix columns : no hardcoded height

            # Gap significance display
            gap_disp = {"high": "HIGH", "medium": "MED", "low": "LOW", "none": ": "}.get(
                sig.gap_significance, ": ")

            row_data = [
                res.url, sig.category, sig.signal_name,
                _fmt_signal_val(sig.raw_value, sig.signal_id),
                _fmt_signal_val(sig.rendered_value, sig.signal_id),
                gap_disp,
                "✓" if sig.match else "✗",
                sig.severity.upper(),
                getattr(sig, "severity_reason", "") or (sol.severity_reason if sol and hasattr(sol, "severity_reason") else ""),
                sol.diagnosis if sol else "",
                sol.fix if sol else "",
            ]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = _border()
                cell.alignment = Alignment(horizontal="left", vertical="top",
                                           wrap_text=(c in (10, 11)), indent=1)
                if c == 8: # Severity
                    bg2, fg = _status_style(sig.severity)
                    cell.fill = _fill(bg2)
                    cell.font = Font(name="Arial", size=9, bold=True, color=fg)
                elif c == 7: # Match
                    cell.fill = _fill(bg)
                    cell.font = _font(bold=True,
                                      color=C_PASS if sig.match else C_CRIT)
                elif c == 6: # Gap significance
                    gap_colors = {"HIGH": (C_CRIT_BG, C_CRIT_FG),
                                  "MED": (C_WARN_BG, C_WARN_FG),
                                  "LOW": (C_INFO_BG, C_INFO_FG)}
                    if val in gap_colors:
                        gbg, gfg = gap_colors[val]
                        cell.fill = _fill(gbg)
                        cell.font = Font(name="Arial", size=9, bold=True, color=gfg)
                    else:
                        cell.fill = _fill(bg)
                        cell.font = _font(sz=9, color=C_TEXT)
                else:
                    cell.fill = _fill(bg)
                    cell.font = _font(sz=9)
            r += 1

    _widths(ws, {"A": 35, "B": 14, "C": 20, "D": 24, "E": 24,
                 "F": 8, "G": 7, "H": 11, "I": 36, "J": 42, "K": 42})


# ── Sheet 4 : Render Gap Analysis ────────────────────────────────────────────
def _sheet_render_gap(wb, results):
    from openpyxl.formatting.rule import ColorScaleRule
    ws = wb.create_sheet("Render Gap Analysis")
    ws.sheet_view.showGridLines = False

    _banner(ws, "RENDER GAP : CONTENT VOLUME DELTA", "A1", "F1", bg=C_SUBHDR, sz=13)
    _header_row(ws, 2, [
        "URL", "Raw Chars", "Rendered Chars", "Gap Chars", "Gap %", "Risk"
    ], bg=C_SUBHDR)

    for r, res in enumerate(results, 3):
        bg = C_ROW_A if r % 2 == 1 else C_ROW_B
        ws.row_dimensions[r].height = 22

        bt = next((s for s in res.signals if s.signal_id == "body_text"), None)
        def _bt_chars(v):
            if isinstance(v, dict): return v.get("char_count", 0)
            return v or 0
        raw_c = _bt_chars(bt.raw_value) if bt else 0
        rend_c = _bt_chars(bt.rendered_value) if bt else 0
        # Compute as Python values : no Excel formulas (avoids blank in CSV/text exports)
        gap_chars = rend_c - raw_c
        if bt and isinstance(bt.rendered_value, dict) and "gap_pct" in bt.rendered_value:
            gap_p = round(bt.rendered_value["gap_pct"], 1)
        else:
            gap_p = round((gap_chars / rend_c * 100), 1) if rend_c else 0.0
        risk = ("HIGH RISK" if gap_p >= 60 else
                     "MEDIUM RISK" if gap_p >= 30 else "LOW RISK")
        risk_bg = C_CRIT_BG if gap_p >= 60 else (C_WARN_BG if gap_p >= 30 else C_PASS_BG)
        risk_fg = C_CRIT_FG if gap_p >= 60 else (C_WARN_FG if gap_p >= 30 else C_PASS_FG)

        data = [res.url, raw_c, rend_c, gap_chars, gap_p / 100, risk]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _border()
            cell.alignment = _align("left" if c == 1 else "center")
            if c == 6:
                cell.fill = _fill(risk_bg)
                cell.font = Font(name="Arial", size=9, bold=True, color=risk_fg)
            else:
                cell.fill = _fill(bg)
                cell.font = _font(bold=(c == 1))
            if c == 5: cell.number_format = "0.0%"
            if c in (2, 3, 4): cell.number_format = "#,##0"

    # Color scale on Gap % column (E) : green at 0%, amber at 30%, red at 60%+
    n = len(results)
    if n > 0:
        gap_range = f"E3:E{2 + n}"
        ws.conditional_formatting.add(gap_range, ColorScaleRule(
            start_type="num", start_value=0, start_color="C6EFCE",
            mid_type="num", mid_value=0.30, mid_color="FFEB9C",
            end_type="num", end_value=0.60, end_color="FFC7CE",
        ))

    # Bar chart
    if n > 0:
        chart = BarChart()
        chart.type = "col"; chart.style = 10
        chart.title = "Raw vs Rendered Content Volume"
        chart.y_axis.title = "Characters"; chart.x_axis.title = "URL"
        chart.width = 22; chart.height = 12
        chart.add_data(Reference(ws, min_col=2, max_col=2, min_row=2, max_row=2+n), titles_from_data=True)
        chart.add_data(Reference(ws, min_col=3, max_col=3, min_row=2, max_row=2+n), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=2+n))
        ws.add_chart(chart, "A" + str(4 + n))

    _widths(ws, {"A": 42, "B": 14, "C": 16, "D": 13, "E": 9, "F": 13})


# ── Sheet 5 : Bot Access Matrix ───────────────────────────────────────────────
def _sheet_bot_matrix(wb, results):
    from config import AI_BOTS
    ws = wb.create_sheet("Bot Access Matrix")
    ws.sheet_view.showGridLines = False

    _banner(ws, "BOT ACCESS MATRIX : robots.txt & llms.txt", "A1",
            get_column_letter(2 + len(AI_BOTS)) + "1", bg=C_SUBHDR, sz=13)
    _header_row(ws, 2, ["Domain"] + AI_BOTS + ["llms.txt"], bg=C_SUBHDR)

    for r, res in enumerate(results, 3):
        from urllib.parse import urlparse
        domain = urlparse(res.url).netloc
        bg = C_ROW_A if r % 2 == 1 else C_ROW_B
        ws.row_dimensions[r].height = 22

        row_data = [domain]
        for bot in AI_BOTS:
            status = res.bot_access.get(bot, "not_mentioned")
            row_data.append(status.upper())
        row_data.append("✓ Present" if res.llms_txt_status == "present" else "✗ Missing")

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _border()
            cell.alignment = _align("center" if c > 1 else "left")
            if val in ("ALLOW", "✓ Present"):
                cell.fill = _fill(C_PASS_BG)
                cell.font = Font(name="Arial", size=9, bold=True, color=C_PASS_FG)
            elif val in ("BLOCK",):
                cell.fill = _fill(C_CRIT_BG)
                cell.font = Font(name="Arial", size=9, bold=True, color=C_CRIT_FG)
            elif val == "✗ Missing":
                cell.fill = _fill(C_WARN_BG)
                cell.font = Font(name="Arial", size=9, bold=True, color=C_WARN_FG)
            else:
                cell.fill = _fill(bg)
                cell.font = _font(bold=(c == 1))

    bw = {"A": 32}
    for i, _ in enumerate(AI_BOTS, 2):
        bw[get_column_letter(i)] = 14
    bw[get_column_letter(2 + len(AI_BOTS))] = 14
    _widths(ws, bw)


# ── Sheet 6 : JS-Gated Content ────────────────────────────────────────────────
def _sheet_js_gated(wb, results):
    """
    Lists every JS-gated content item:
      1. Text block samples from the Python extractor (js_gated_text_samples)
      2. AI-detected JS-gating from solutions : any WARNING/CRITICAL solution
         whose signal_id is a known JS-gated signal OR whose diagnosis
         explicitly mentions JavaScript rendering.
    """
    # Signals that are inherently about JS-gated content
    JS_GATED_SIGNAL_IDS = {"body_text", "h1", "h2", "h3", "h4", "h5",
                           "internal_links", "json_ld", "og_title"}
    JS_KEYWORDS = ("javascript", "js-gated", "dynamically", "client-side",
                   "rendered dom", "injected by", "benefits-root", "benifits")

    ws = wb.create_sheet("JS-Gated Content")
    ws.sheet_view.showGridLines = False

    _banner(ws, "JS-GATED CONTENT : PRESENT IN RENDERED DOM, ABSENT FROM RAW HTML",
            "A1", "E1", bg=C_SUBHDR, sz=12)
    _header_row(ws, 2, ["URL", "Signal", "Source", "What is JS-gated", "Specific content / evidence"], bg=C_SUBHDR)

    r = 3
    seen_sol_ids = set() # avoid duplicate rows per signal

    for res in results:
        # ── Part 1: Python extractor text block samples ──────────────────────
        bt = next((s for s in res.signals if s.signal_id == "body_text"), None)
        if bt and isinstance(bt.rendered_value, dict):
            samples = bt.rendered_value.get("js_gated_text_samples", [])
            for sample in samples:
                bg = C_ROW_A if r % 2 == 1 else C_ROW_B
                ws.row_dimensions[r].height = None
                for c, val in enumerate(
                    [res.url, "Body Text Volume", "Python extractor (DOM diff)", "Text block absent from raw HTML", sample], 1
                ):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = _border()
                    cell.fill = _fill(bg)
                    cell.font = _font(sz=9)
                    cell.alignment = Alignment(horizontal="left", vertical="top",
                                               wrap_text=(c in (4, 5)), indent=1)
                r += 1

        # ── Part 2: AI-detected JS-gating from solutions ─────────────────────
        for sol in res.solutions:
            if sol.severity not in ("warning", "critical"):
                continue

            key = (res.url, sol.signal_id)
            if key in seen_sol_ids:
                continue

            diag_lower = (sol.diagnosis or "").lower()
            obs_lower = (sol.observed_in_raw or "").lower() + (sol.observed_in_rendered or "").lower()

            is_js_gated_signal = sol.signal_id in JS_GATED_SIGNAL_IDS
            mentions_js = any(kw in diag_lower or kw in obs_lower for kw in JS_KEYWORDS)

            if not (is_js_gated_signal or mentions_js):
                continue

            seen_sol_ids.add(key)
            bg = C_ROW_A if r % 2 == 1 else C_ROW_B
            ws.row_dimensions[r].height = None

            label = SIGNAL_LABEL_MAP.get(sol.signal_id, sol.signal_id.replace("_", " ").title())
            what = sol.observed_in_raw or "Empty / absent in raw HTML"
            evidence = sol.observed_in_rendered or sol.diagnosis or ""

            for c, val in enumerate([res.url, label, "AI reading (Gemini)", what, evidence], 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = _border()
                cell.fill = _fill(bg)
                cell.alignment = Alignment(horizontal="left", vertical="top",
                                           wrap_text=(c in (4, 5)), indent=1)
                if c == 3:
                    cell.font = Font(name="Arial", size=9, color=C_INFO, italic=True)
                elif c == 2:
                    bg2, fg = _status_style(sol.severity)
                    cell.fill = _fill(bg2)
                    cell.font = Font(name="Arial", size=9, bold=True, color=fg)
                else:
                    cell.fill = _fill(bg)
                    cell.font = _font(sz=9)
            r += 1

    if r == 3:
        ws.cell(row=3, column=1,
                value="No JS-gated content detected.").font = _font(color="888888")

    _widths(ws, {"A": 35, "B": 26, "C": 22, "D": 35, "E": 70})


# ── Signal label map for xlsx (mirrors template SIGNAL_LABELS) ────────────────
SIGNAL_LABEL_MAP = {
    "title": "Title Tag",
    "meta_description": "Meta Description",
    "meta_robots": "Meta Robots",
    "x_robots_header": "X-Robots-Tag Header",
    "canonical": "Canonical Tag",
    "h1": "H1 Headings",
    "h2": "H2 Headings hidden by JavaScript",
    "h3": "H3 Headings hidden by JavaScript",
    "body_text": "Body Text JS-gated",
    "internal_links": "Internal Links",
    "json_ld": "JSON-LD Schema",
    "og_title": "OG Title",
    "images": "Images",
    "images_missing_alt": "Images : Missing Alt Text",
}


# ── Sheet 7 : Missing Alt Images ─────────────────────────────────────────────
def _sheet_missing_alt(wb, results):
    """
    Full list of images with empty/missing alt attributes.
    Includes an AI-suggested alt text column (sourced from solution code_fix)
    clearly labelled as AI-generated : must be reviewed before use.
    """
    import re

    ws = wb.create_sheet("Missing Alt Images")
    ws.sheet_view.showGridLines = False

    _banner(ws, "IMAGES WITH MISSING ALT ATTRIBUTES",
            "A1", "E1", bg=C_SUBHDR, sz=12)

    # Warning note row
    ws.merge_cells("A2:E2")
    note = ws["A2"]
    note.value = (
        "⚠ Column E contains AI-suggested alt text : inferred from filenames/context by Gemini. "
        "These values DO NOT exist on the live page. Review every suggestion and rewrite before deploying."
    )
    note.font = Font(name="Arial", size=9, bold=True, color=C_WARN_FG)
    note.fill = _fill(C_WARN_BG)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    note.border = _border()
    ws.row_dimensions[2].height = 32

    _header_row(ws, 3, [
        "URL", "#", "Image src (exact from HTML)",
        "Current alt value", "AI-suggested alt text (review before use)"
    ], bg=C_SUBHDR)

    r = 4
    for res in results:
        img_sig = next((s for s in res.signals if s.signal_id == "images"), None)
        if not img_sig:
            continue
        rv = img_sig.rendered_value
        if not isinstance(rv, dict):
            continue
        missing = rv.get("missing_alt", [])
        if not missing:
            continue

        # Build src → suggested alt map from solution code_fix/code_block
        # The solution for "images" may contain HTML with alt="..." filled by Gemini
        src_to_suggestion = {}
        img_sol = next(
            (sol for sol in res.solutions if sol.signal_id in ("images", "images_missing_alt")),
            None
        )
        if img_sol:
            code = img_sol.code_block or img_sol.code_fix or ""
            # Extract src + alt pairs from the code block
            pairs = re.findall(r'src=["\']([^"\']+)["\'][^>]*alt=["\']([^"\']*)["\']|alt=["\']([^"\']*)["\'][^>]*src=["\']([^"\']+)["\']', code)
            for m in pairs:
                # Group (src, alt, alt, src) depending on order
                src = (m[0] or m[3]).strip()
                alt = (m[1] or m[2]).strip()
                if src and alt and alt not in ("[DESCRIBE: write your own alt text here]", ""):
                    src_to_suggestion[src] = alt
                elif src and alt in ("[DESCRIBE: write your own alt text here]", ""):
                    src_to_suggestion[src] = "" # placeholder : nothing useful

        for idx, img in enumerate(missing, 1):
            bg = C_ROW_A if r % 2 == 1 else C_ROW_B
            ws.row_dimensions[r].height = 22
            src = img.get("src", "") if isinstance(img, dict) else str(img)
            suggestion = src_to_suggestion.get(src, "")

            row_data = [res.url, idx, src, '"" (empty)', suggestion or ": "]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = _border()
                cell.fill = _fill(bg)
                cell.alignment = _align("left" if c != 2 else "center")
                if c == 3:
                    cell.font = Font(name="Arial", size=9, color="4472C4")
                elif c == 4:
                    cell.font = Font(name="Arial", size=9, color=C_CRIT_FG, bold=True)
                elif c == 5:
                    if suggestion:
                        # Clearly colour-coded as AI suggestion
                        cell.fill = _fill(C_INFO_BG)
                        cell.font = Font(name="Arial", size=9, color=C_INFO_FG, italic=True)
                    else:
                        cell.font = Font(name="Arial", size=9, color="888888", italic=True)
                else:
                    cell.font = _font(sz=9)
            r += 1

    if r == 4:
        ws.cell(row=4, column=1,
                value="No images with missing alt attributes detected.").font = _font(color="888888")

    _widths(ws, {"A": 35, "B": 5, "C": 70, "D": 16, "E": 45})

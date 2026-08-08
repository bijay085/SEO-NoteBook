"""Append the measured data tabs to the branded workbook.

`report_kit.render_xlsx` renders the AUTHORED half (Overview + one tab per
section, every finding as Issue·Evidence·Solution·Execution). A log audit also
has to ship the raw measured half : the tabs a client actually filters and
sorts: which URLs, which bots, which day. That is what this adds.

Separate module on purpose: `report_kit.py` is shared byte-identical across
every seo-* skill, so log-specific sheets must never be added to it.

Tabs appended, and only when the data exists (no hollow skeletons):
  Health Scorecard · Traffic Breakdown · URL Detail · Bot Crawl Detail ·
  Crawl Trend · Log Sources · Crawl Coverage · Decision Guide

append_data_tabs(xlsx_path, analysis_json_path) -> [tab names]
"""
import json

YL, BK = "FFF5C518", "FF0A0A0A"
GREEN, YELLOW, RED = "FFDCFCE7", "FFFEF9C3", "FFFEE2E2"

# Interpretation reference, so a reader who is not an SEO can act on the
# findings tabs without a call.
STATUS_GUIDE = [
    ["200", "OK : page served correctly", "No", "OK", "None needed", "Pages render correctly"],
    ["301", "Moved Permanently", "Yes", "P2 - High", "Update internal links + sitemap to the final URL", "301 hits trend to zero"],
    ["302", "Temporary Redirect", "Yes", "P2 - High", "Change to 301 if permanent; update source links", "curl -Iv shows 301"],
    ["304", "Not Modified (cache revalidation)", "No", "OK", "None. Efficient caching.", "Normal, especially on robots.txt"],
    ["400", "Bad Request : malformed URL", "Yes", "P2 - High", "Fix the malformed link generating it", "URL leaves the bot logs"],
    ["403", "Forbidden : access blocked", "Depends", "P2 - High", "Confirm no important pages are blocked", "No money pages return 403"],
    ["404", "Not Found : page missing", "YES", "P1 - Critical", "301 or 410; fix internal links; update sitemap", "Zero 404s on indexed URLs"],
    ["410", "Gone : intentionally deleted", "No", "OK", "Correct signal for permanent removal", "Crawl hits fall over 2-4 weeks"],
    ["429", "Too Many Requests : rate limited", "Depends", "P2 - High", "Ensure Googlebot is not being rate-limited", "Googlebot absent from 429s"],
    ["500", "Internal Server Error", "YES", "P1 - Critical", "Fix now. Check PHP/server error logs.", "URL returns 200; Crawl Stats normal"],
    ["502", "Bad Gateway : upstream failure", "YES", "P1 - Critical", "Check PHP-FPM / Node process and the proxy", "curl -I returns 200"],
    ["503", "Service Unavailable", "If persistent", "P1 - Critical", "Add Retry-After; fix the capacity issue", "Alert if over 1 hour"],
    ["504", "Gateway Timeout : upstream slow", "YES", "P1 - Critical", "Fix the slow DB/PHP path or raise the timeout", "Response time <500ms"],
]

SEGMENT_GUIDE = [
    ["Search Bots", "Googlebot, Bingbot, Yandex : verified by IP range", "Counts directly toward crawl budget", "Monitor what they crawl and the status codes they see."],
    ["SEO Crawlers", "Ahrefs, Semrush, Screaming Frog, Sitebulb", "Server load + bandwidth cost", "Throttle in robots.txt if they crawl too often."],
    ["Infrastructure", "WP self-pings, cache preloaders, uptime monitors", "Inflates log counts; server load", "Tune the frequency; schedule off-peak."],
    ["Generic / AI Bots", "curl, wget, python-requests, GPTBot, ClaudeBot, PerplexityBot", "Minimal search impact; some load", "Decide deliberately whether to allow AI fetches."],
    ["Suspicious", "High-volume IPs and impossible-version UAs, after the admin-URL gate", "Bandwidth + CPU cost; scraping risk", "Block at the CDN/WAF : verify the IP is a datacenter first."],
    ["Human", "Real browser users, including logged-in admins", "Not crawl budget", "Analyse in GA4/GSC; optimise UX and conversion."],
]


def append_data_tabs(xlsx_path, analysis_json_path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with open(analysis_json_path, "r", encoding="utf-8") as fh:
        data = json.load(fh)

    wb = load_workbook(xlsx_path)
    hfill = PatternFill("solid", fgColor=BK)
    hfont = Font(name="Arial", bold=True, color="FFFFFFFF")
    base = Font(name="Arial")
    bold = Font(name="Arial", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    added = []

    def sheet(title, cols, rows, widths=None, note=None):
        if not rows:
            return None
        ws = wb.create_sheet(title[:31])
        r = 1
        if note:
            ws.cell(1, 1, note).font = Font(name="Arial", italic=True, size=9)
            ws.cell(1, 1).alignment = wrap
            r = 3
        for j, c in enumerate(cols, 1):
            cell = ws.cell(r, j, c)
            cell.font = hfont
            cell.fill = hfill
            ws.column_dimensions[get_column_letter(j)].width = (
                widths[j - 1] if widths and j - 1 < len(widths) else 18)
        for i, row in enumerate(rows):
            for j, v in enumerate(row, 1):
                cell = ws.cell(r + 1 + i, j, v)
                cell.font = base
                cell.alignment = wrap
        ws.freeze_panes = ws.cell(r + 1, 1).coordinate
        added.append(ws.title)
        return ws

    m = data.get("meta", {})

    # ── Issue Analysis : THE deterministic findings sheet ──────────────────
    # 16 columns, every finding, produced with no authoring step. This is the
    # sheet the operator actually works from: what is wrong, how confident we
    # are and why, what caused it, what it costs, the literal fix, the effort,
    # and how to prove it is fixed. The authored per-dimension tabs are an
    # interpretation layer ON TOP of this : never a replacement for it.
    findings = data.get("findings") or []
    if findings:
        ws = wb.create_sheet("Issue Analysis", 1 if "Overview" in wb.sheetnames else 0)
        hdrs = ["#", "Category", "Issue", "Detail", "Affected URL / Scope", "Segment",
                "Status", "Hits", "Urgency", "Confidence", "Evidence", "Root Cause",
                "Business Impact", "Step-by-Step Action", "Effort Estimate",
                "Verification Steps"]
        widths = [4, 16, 24, 38, 42, 14, 8, 8, 14, 11, 40, 34, 30, 56, 18, 44]
        p1 = sum(1 for f in findings if f["urgency"] == "P1 - Critical")
        p2 = sum(1 for f in findings if f["urgency"] == "P2 - High")
        p3 = sum(1 for f in findings if f["urgency"] == "P3 - Monitor")
        ws.cell(1, 1, f'{m.get("site", "")} : SEO Issue Analysis').font = Font(
            name="Arial", bold=True, size=14)
        ws.cell(2, 1, f'{m.get("date_range", {}).get("start", "?")} to '
                      f'{m.get("date_range", {}).get("end", "?")} | '
                      f'{m.get("total_requests", 0):,} requests | '
                      f'{len(findings)} findings: P1 {p1} · P2 {p2} · P3 {p3} | '
                      f'parse rate {m.get("parse_rate_pct", 0)}%').font = Font(
            name="Arial", size=9, italic=True)
        for j, h in enumerate(hdrs, 1):
            cell = ws.cell(4, j, h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
        urg_fill = {"P1 - Critical": RED, "P2 - High": YELLOW, "P3 - Monitor": "FFDBEAFE"}
        conf_fill = {"High": GREEN, "Medium": YELLOW, "Low": RED}
        for i, f in enumerate(findings, 1):
            r = 4 + i
            vals = [i, f["category"], f["issue"], f["detail"], f["url"], f["segment"],
                    f["status_code"], f["hits"], f["urgency"], f["confidence"],
                    f["evidence"], f["root_cause"], f["impact"], f["action"],
                    f["effort"], f["verify"]]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(r, j, v)
                cell.font = base
                cell.alignment = wrap
            ws.cell(r, 9).fill = PatternFill(
                "solid", fgColor=urg_fill.get(f["urgency"], "FFFFFFFF"))
            ws.cell(r, 9).font = bold
            ws.cell(r, 10).fill = PatternFill(
                "solid", fgColor=conf_fill.get(f["confidence"], "FFFFFFFF"))
            ws.row_dimensions[r].height = 78
        ws.freeze_panes = "E5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(hdrs))}{4 + len(findings)}"
        added.append(ws.title)

    # Grouped view : one row per pattern, so a 131-URL 404 problem reads as one
    # line before anyone opens the 131-row detail.
    sheet("Findings Summary",
          ["Worst Urgency", "Category", "Issue", "URLs Affected", "Total Hits",
           "Confidence Mix", "Worst-Offender URLs"],
          [[g["worst_urgency"], g["category"], g["issue"], g["count"], g["total_hits"],
            ", ".join(f'{k}x{v}' for k, v in g["confidence"].items()),
            "\n".join(f'{t["url"]} ({t["hits"]} hits)' for t in g["top"])]
           for g in (data.get("findings_summary") or [])],
          [15, 18, 30, 13, 12, 22, 62])

    # Health Scorecard : colour the verdict column.
    health = data.get("health") or []
    ws = sheet("Health Scorecard",
               ["Metric", "Your Value", "Score", "Healthy", "Watch", "Urgent"],
               [[h["metric"], h["value"], h["score"], h["healthy"], h["watch"],
                 h["urgent"]] for h in health],
               [26, 14, 10, 14, 14, 14])
    if ws:
        for i, h in enumerate(health):
            cell = ws.cell(2 + i, 3)
            cell.fill = PatternFill(
                "solid", fgColor={"Green": GREEN, "Yellow": YELLOW,
                                  "Red": RED}.get(h["score"], "FFFFFFFF"))
            cell.font = bold

    ws = sheet("Traffic Breakdown",
               ["Segment", "Requests", "% of Total"],
               [[s["segment"], s["requests"], f'{s["pct"]}%']
                for s in data.get("segments", [])],
               [24, 14, 14],
               note=f'{m.get("site", "")} · {m.get("total_requests", 0):,} requests · '
                    f'{m.get("verified_bot_coverage", "")}')
    agents = data.get("agents") or []
    if ws and agents:
        start = ws.max_row + 3
        ws.cell(start, 1, "TOP AGENTS").font = bold
        for j, c in enumerate(["Agent", "Class", "Requests", "% of Total",
                               "Unique URLs", "Top Status"], 1):
            cell = ws.cell(start + 1, j, c)
            cell.font = hfont
            cell.fill = hfill
        for i, a in enumerate(agents):
            for j, v in enumerate([a["agent"], a["class"], a["requests"],
                                   f'{a["pct"]}%', a["unique_urls"],
                                   a["top_status"]], 1):
                ws.cell(start + 2 + i, j, v).font = base
        ws.column_dimensions["A"].width = 40

    sheet("URL Detail",
          ["URL", "Total Hits", "Search Bot", "Infra", "Suspicious", "Human",
           "Status", "Avg Bytes", "Content Type", "Last Seen"],
          [[u["url"], u["total_hits"], u["search_bot"], u["infra"], u["suspicious"],
            u["human"], u["status"], u["avg_bytes"], u["content_type"],
            u["last_seen"]] for u in data.get("url_detail", [])],
          [58, 11, 11, 9, 11, 9, 8, 11, 14, 22])

    sheet("Bot Crawl Detail",
          ["URL", "Search Engine", "Bot", "Hits", "Status", "Content Type",
           "Avg Bytes", "Last Seen"],
          [[b["url"], b["search_engine"], b["bot"], b["hits"], b["status"],
            b["content_type"], b["avg_bytes"], b["last_seen"]]
           for b in data.get("bot_crawl", [])],
          [58, 14, 22, 9, 8, 14, 11, 22],
          note="Search-engine crawler requests only : verified by IP range where the "
               "engine publishes one, otherwise identified by User-Agent.")

    trend = data.get("crawl_trend") or []
    if trend:
        keys = [k for k in trend[0] if k not in ("date", "total")]
        sheet("Crawl Trend", ["Date", "Total"] + keys,
              [[t["date"], t["total"]] + [t.get(k, 0) for k in keys] for t in trend],
              [14, 12] + [16] * len(keys))

    sheet("Log Sources",
          ["Log Role", "Files", "Rows Parsed", "Unique URLs",
           "Used In Main Totals", "Purpose", "Filenames"],
          [[s["role"], s["files"], s["rows"], s["unique_urls"],
            "Yes" if s["used_in_main_totals"] else "No", s["purpose"],
            ", ".join(s["filenames"])] for s in data.get("log_sources", [])],
          [20, 8, 13, 13, 18, 46, 46],
          note=f'Routing decision: {m.get("routing_reason", "")}')

    cross = data.get("crossref") or {}
    cross_rows = []
    if "sitemap" in cross:
        c = cross["sitemap"]
        cross_rows.append(["Sitemap",
                           f'{c["never_crawled"]} of {c["sitemap_urls"]} sitemap URLs '
                           f'never crawled', c["note"], ""])
        cross_rows += [["Sitemap", "Never crawled in this window", "", u]
                       for u in c.get("sample", [])]
    if "gsc" in cross:
        c = cross["gsc"]
        cross_rows.append(["GSC",
                           f'{c["impressions_but_no_crawl"]} of {c["gsc_pages"]} pages '
                           f'have impressions but no crawl', c["note"], ""])
        cross_rows += [["GSC", f'{s["impressions"]} impressions, no crawl', "", s["url"]]
                       for s in c.get("sample", [])]
    sheet("Crawl Coverage", ["Source", "Finding", "Caveat", "URL"], cross_rows,
          [12, 46, 52, 58])

    ws = sheet("Decision Guide",
               ["Code", "Meaning", "SEO Problem?", "Priority", "Action", "Verify"],
               STATUS_GUIDE, [8, 34, 15, 15, 46, 34])
    if ws:
        for i, row in enumerate(STATUS_GUIDE):
            cell = ws.cell(2 + i, 4)
            if "P1" in row[3]:
                cell.fill = PatternFill("solid", fgColor=RED)
            elif "P2" in row[3]:
                cell.fill = PatternFill("solid", fgColor=YELLOW)
            else:
                cell.fill = PatternFill("solid", fgColor=GREEN)
        start = ws.max_row + 3
        ws.cell(start, 1, "TRAFFIC SEGMENT REFERENCE").font = bold
        for j, c in enumerate(["Segment", "What It Is", "Crawl Budget Impact",
                               "Action"], 1):
            cell = ws.cell(start + 1, j, c)
            cell.font = hfont
            cell.fill = hfill
        for i, row in enumerate(SEGMENT_GUIDE):
            for j, v in enumerate(row, 1):
                cell = ws.cell(start + 2 + i, j, v)
                cell.font = base
                cell.alignment = wrap

    wb.save(xlsx_path)
    return added

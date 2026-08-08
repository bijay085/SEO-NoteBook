#!/usr/bin/env python3
"""branded XLSX helpers for the Initial Analysis master workbook.

openpyxl primitives so the team's execute-from workbook matches the HTML reports:
black header bands, yellow accents, colored sheet tabs, frozen + styled header rows,
alternating row fill, status-color left-border accents, gridlines off, Arial. The
SHEET CONTENT is client-specific (you fill it per run); this lib supplies the look.

    import workbook_lib as W
    wb = W.new_book()
    ws = W.add_sheet(wb, "Overview", W.YELLOW)
    nxt = W.header_band(ws, "Acme", "Initial Analysis", ncols=6)
    hr = W.head_row(ws, nxt + 1, ["Metric", "Value"])
    W.write_rows(ws, hr, [["Entity", "Food Photography"]])
    W.widths(ws, [26, 60]); W.save(wb, "00-Acme-Master-Workbook.xlsx")
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

YELLOW = "F5C518"; BLACK = "0A0A0A"; DARK = "1A1A1A"; GREEN = "2ECC71"; RED = "E74C3C"
ORANGE = "E67E22"; MUTED = "888888"; BG = "F7F7F7"; WHITE = "FFFFFF"; TEXT = "1C1C1C"; ALT = "F2F2F2"
FONT = "Arial"
# map a status/verdict string -> accent color for the left border of a data row
STATUS = {"done": GREEN, "in-progress": ORANGE, "blocked": RED, "todo": MUTED,
          "confirmed": GREEN, "false": RED, "needs-data": ORANGE, "pending": ORANGE}

def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def _side(hex_, style="thin"):
    return Side(style=style, color=hex_)

def new_book():
    wb = Workbook(); wb.remove(wb.active); return wb

def add_sheet(wb, title, tab_color=YELLOW):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color
    return ws

def header_band(ws, client, title, ncols=6, row=1):
    """Black band with client · title; a thin yellow accent row beneath. Returns
    the next free row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"{client} · {title}")
    c.fill = _fill(BLACK); c.font = Font(name=FONT, bold=True, size=14, color=YELLOW)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 30
    for col in range(1, ncols + 1):
        ws.cell(row=row + 1, column=col).fill = _fill(YELLOW)
    ws.row_dimensions[row + 1].height = 4
    return row + 2

def head_row(ws, row, headers, freeze=True):
    """Styled header row (black bg, yellow bold text); freezes panes below it."""
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _fill(BLACK); c.font = Font(name=FONT, bold=True, size=10, color=YELLOW)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)
    return row + 1

def write_rows(ws, start_row, rows, status_col=None):
    """Data rows: alternating fill, wrapped, Arial. If status_col is given, the row's
    first cell gets a thick left border colored by STATUS[row[status_col]]."""
    r = start_row
    for i, row in enumerate(rows):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name=FONT, size=10, color=TEXT)
            c.alignment = Alignment(vertical="top", horizontal="left", indent=1, wrap_text=True)
            if i % 2:
                c.fill = _fill(ALT)
        if status_col is not None and status_col < len(row):
            color = STATUS.get(str(row[status_col]).strip().lower())
            if color:
                ws.cell(row=r, column=1).border = Border(left=_side(color, "thick"))
        r += 1
    return r

def stat_cards(ws, row, cards):
    """cards = [(label, value), ...] -> a row of dark KPI cells, yellow values."""
    for col, (label, value) in enumerate(cards, 1):
        v = ws.cell(row=row, column=col, value=value)
        v.fill = _fill(DARK); v.font = Font(name=FONT, bold=True, size=16, color=YELLOW)
        v.alignment = Alignment(vertical="center", horizontal="center")
        l = ws.cell(row=row + 1, column=col, value=label)
        l.fill = _fill(DARK); l.font = Font(name=FONT, size=9, color=MUTED)
        l.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[row].height = 26
    return row + 2

def widths(ws, cols):
    for i, w in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def save(wb, path):
    wb.save(path); print(f"Wrote {path} ({len(wb.sheetnames)} sheets)")

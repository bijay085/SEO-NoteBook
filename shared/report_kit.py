"""Canonical SEO deliverable renderer (pure stdlib for HTML).

One copy lives at repo-root `shared/report_kit.py`. Skill scripts add that
folder to sys.path. Do not copy this file into individual skill folders.

render_html(report) -> str standalone branded HTML (jump-nav, <details>
                                  accordions, Issue·Evidence·Solution·Execution
                                  cards, inline SVG charts).
render_xlsx(report, path) -> None branded workbook at parity (needs openpyxl).

report schema:
  {
    "title": "Deep Audit", "client": "Acme", "period": "Q3 2026",
    "subtitle": "...", "output_dir": "./Deliverable",
    "sections": [
      { "id": "perf", "title": "Measured Performance", "intro": "...",
        "chart": {"type": "bars"|"hbars", "title": "...", "unit": "",
                  "data": [["LCP", 2.1], ["CLS", 0.03]]},
        "table": {"cols": ["Page","Issue"], "rows": [["/x","..."]]},
        "findings": [
          {"issue": "...", "sev": "critical|high|medium|low|good|info",
           "evidence": "...", "solution": "...", "execution": "...",
           "effort": "M", "priority": "P0"} ] } ]
  }
Every finding MUST carry issue+evidence+solution+execution : that is the quality bar.
"""
import html as _h

BRAND = {
    "yellow": "#F5C518", "black": "#0A0A0A", "dark": "#1A1A1A",
    "paper": "#FFFFFF", "ink": "#14140F", "muted": "#6B6B63", "line": "#E7E7DF",
    "green": "#2ECC71", "red": "#E74C3C", "orange": "#E67E22", "blue": "#3498DB",
}
SEV = {
    "critical": ("Critical", BRAND["red"]),
    "high": ("High", BRAND["orange"]),
    "medium": ("Medium", BRAND["blue"]),
    "low": ("Low", BRAND["muted"]),
    "good": ("Good", BRAND["green"]),
    "info": ("Info", BRAND["muted"]),
}
SEV_ORDER = ["critical", "high", "medium", "low", "good", "info"]


def esc(x):
    return _h.escape("" if x is None else str(x))


def _sevkey(s):
    k = (s or "info").strip().lower()
    return k if k in SEV else "info"


def _num(v):
    f = float(v)
    return str(int(f)) if f == int(f) else f"{f:.2f}".rstrip("0").rstrip(".")


def svg_bars(chart):
    data = [(str(l), float(v)) for l, v in (chart.get("data") or []) if v is not None]
    if not data:
        return ""
    title = chart.get("title", "")
    unit = chart.get("unit", "")
    horizontal = chart.get("type") == "hbars"
    mx = max(v for _, v in data) or 1.0
    Y, R = BRAND["yellow"], BRAND["black"]
    if horizontal:
        rowh, labelw, w, pad = 26, 170, 600, 6
        barw = w - labelw - 80
        h = pad * 2 + rowh * len(data)
        p = []
        for i, (l, v) in enumerate(data):
            y = pad + i * rowh
            bw = max(2.0, (v / mx) * barw)
            p.append(f'<text x="0" y="{y+17}" class="c-lbl">{esc(l)}</text>')
            p.append(f'<rect x="{labelw}" y="{y+5}" width="{bw:.1f}" height="15" rx="3" '
                     f'fill="{Y}" stroke="{R}" stroke-width="0.6"/>')
            p.append(f'<text x="{labelw+bw+6:.1f}" y="{y+17}" class="c-val">{_num(v)}{esc(unit)}</text>')
        svg = (f'<svg viewBox="0 0 {w} {h}" class="c-svg" preserveAspectRatio="xMinYMin meet" '
               f'role="img">{"".join(p)}</svg>')
    else:
        n, w, h, gap, top = len(data), 680, 240, 12, 16
        base = h - 28
        bw = (w - gap * (n + 1)) / n
        p = [f'<line x1="0" y1="{base}" x2="{w}" y2="{base}" stroke="{BRAND["line"]}"/>']
        for i, (l, v) in enumerate(data):
            x = gap + i * (bw + gap)
            bh = max(2.0, (v / mx) * (base - top))
            y = base - bh
            p.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bw:.1f}" height="{bh:.1f}" rx="3" '
                     f'fill="{Y}" stroke="{R}" stroke-width="0.6"/>')
            p.append(f'<text x="{x+bw/2:.1f}" y="{y-4:.1f}" class="c-val" text-anchor="middle">{_num(v)}{esc(unit)}</text>')
            p.append(f'<text x="{x+bw/2:.1f}" y="{base+15:.1f}" class="c-lbl" text-anchor="middle">{esc(l)}</text>')
        svg = (f'<svg viewBox="0 0 {w} {h}" class="c-svg" preserveAspectRatio="xMinYMin meet" '
               f'role="img">{"".join(p)}</svg>')
    return f'<figure class="chart"><figcaption>{esc(title)}</figcaption>{svg}</figure>'


def _table(t):
    cols = t.get("cols") or []
    rows = t.get("rows") or []
    if not cols and not rows:
        return ""
    head = "".join(f"<th>{esc(c)}</th>" for c in cols)
    body = "".join("<tr>" + "".join(f"<td>{esc(c)}</td>" for c in r) + "</tr>" for r in rows)
    return f'<div class="tw"><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>'


def _finding(f):
    key = _sevkey(f.get("sev"))
    label, color = SEV[key]
    openattr = " open" if key in ("critical", "high") else ""
    meta = []
    if f.get("effort"):
        meta.append(f'Effort: {esc(f["effort"])}')
    if f.get("priority"):
        meta.append(f'Priority: {esc(f["priority"])}')
    metahtml = f'<div class="f-meta">{" · ".join(meta)}</div>' if meta else ""

    def block(lbl, val, cls=""):
        if not val:
            return ""
        if cls == "exec":
            return f'<div class="f-row"><span class="f-k">{lbl}</span><pre class="f-exec">{esc(val)}</pre></div>'
        return f'<div class="f-row"><span class="f-k">{lbl}</span><div class="f-v">{esc(val)}</div></div>'

    return (
        f'<details class="f sev-{key}"{openattr}>'
        f'<summary><span class="chip" style="background:{color}">{label}</span>'
        f'<span class="f-issue">{esc(f.get("issue"))}</span></summary>'
        f'<div class="f-body">'
        f'{block("Evidence", f.get("evidence"))}'
        f'{block("Solution", f.get("solution"))}'
        f'{block("Execution", f.get("execution"), "exec")}'
        f'{metahtml}'
        f'</div></details>'
    )


def _section(s):
    fnd = s.get("findings") or []
    inner = "".join(_finding(f) for f in fnd) or \
        '<p class="empty">No findings recorded yet : populate this section in <code>report_data.py</code>.</p>'
    intro = f'<p class="s-intro">{esc(s["intro"])}</p>' if s.get("intro") else ""
    chart = svg_bars(s["chart"]) if s.get("chart") else ""
    table = _table(s["table"]) if s.get("table") else ""
    return (
        f'<section id="sec-{esc(s["id"])}" class="sec">'
        f'<h2>{esc(s["title"])}</h2>{intro}{chart}{table}'
        f'<div class="findings">{inner}</div></section>'
    )


def _sev_counts(sections):
    c = {k: 0 for k in SEV_ORDER}
    for s in sections:
        for f in (s.get("findings") or []):
            c[_sevkey(f.get("sev"))] += 1
    return c


def render_html(report):
    sections = report.get("sections") or []
    client = report.get("client", "Client")
    title = report.get("title", "Audit")
    period = report.get("period", "")
    header = f'{esc(client)} · {esc(title)}' + (f' : {esc(period)}' if period else "")
    counts = _sev_counts(sections)
    total = sum(counts.values())

    nav = "".join(f'<a href="#sec-{esc(s["id"])}">{esc(s["title"])}</a>' for s in sections)

    cards = []
    for s in sections:
        fnd = s.get("findings") or []
        top = "info"
        for k in SEV_ORDER:
            if any(_sevkey(f.get("sev")) == k for f in fnd):
                top = k
                break
        cards.append(
            f'<a class="card" href="#sec-{esc(s["id"])}">'
            f'<span class="dot" style="background:{SEV[top][1]}"></span>'
            f'<span class="card-t">{esc(s["title"])}</span>'
            f'<span class="card-n">{len(fnd)} finding{"s" if len(fnd)!=1 else ""}</span></a>'
        )
    grid = f'<div class="grid">{"".join(cards)}</div>'

    sev_chart = svg_bars({
        "type": "bars", "title": f"Findings by severity ({total} total)",
        "data": [(SEV[k][0], counts[k]) for k in SEV_ORDER if counts[k]],
    }) if total else ""

    body = "".join(_section(s) for s in sections)
    sub = f'<p class="sub">{esc(report["subtitle"])}</p>' if report.get("subtitle") else ""

    return f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(client)} : {esc(title)}</title>
<style>
:root{{--yl:{BRAND['yellow']};--bk:{BRAND['black']};--dk:{BRAND['dark']};
--ink:{BRAND['ink']};--mut:{BRAND['muted']};--ln:{BRAND['line']};}}
*{{box-sizing:border-box}}
body{{margin:0;font:15px/1.55 Inter,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;
color:var(--ink);background:#FAFAF7}}
a{{color:inherit}}
header.top{{background:var(--bk);color:#fff;padding:26px 22px}}
header.top .brand{{display:inline-block;background:var(--yl);color:var(--bk);font-weight:800;
padding:2px 9px;border-radius:5px;letter-spacing:.5px;font-size:13px}}
header.top h1{{margin:10px 0 0;font-size:26px;font-weight:800}}
header.top .sub{{margin:6px 0 0;color:#c9c9c2;max-width:70ch}}
nav.jump{{position:sticky;top:0;z-index:20;background:var(--dk);border-bottom:2px solid var(--yl);
display:flex;gap:2px;overflow-x:auto;padding:8px 10px;-webkit-overflow-scrolling:touch}}
nav.jump a{{color:#e9e9e2;text-decoration:none;white-space:nowrap;font-size:12.5px;
padding:6px 10px;border-radius:6px}}
nav.jump a:hover{{background:#000;color:var(--yl)}}
main{{max-width:1000px;margin:0 auto;padding:22px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(min(220px,100%),1fr));gap:12px;margin:8px 0 26px}}
.card{{display:flex;flex-direction:column;gap:3px;background:#fff;border:1px solid var(--ln);
border-radius:10px;padding:13px 14px;text-decoration:none;transition:.15s}}
.card:hover{{border-color:var(--bk);box-shadow:0 3px 10px rgba(0,0,0,.07);transform:translateY(-1px)}}
.card .dot{{width:9px;height:9px;border-radius:50%}}
.card-t{{font-weight:700;font-size:14px}} .card-n{{color:var(--mut);font-size:12px}}
.sec{{background:#fff;border:1px solid var(--ln);border-radius:12px;padding:20px 20px 8px;margin:0 0 20px;scroll-margin-top:60px}}
.sec h2{{margin:0 0 4px;font-size:19px;border-left:4px solid var(--yl);padding-left:10px}}
.s-intro{{color:var(--mut);margin:6px 0 14px}}
.empty{{color:var(--mut);font-style:italic}}
figure.chart{{margin:6px 0 16px;padding:12px;border:1px solid var(--ln);border-radius:9px;background:#FCFCF9}}
figure.chart figcaption{{font-weight:700;font-size:13px;margin-bottom:6px}}
.c-svg{{width:100%;height:auto}} .c-lbl{{font-size:11px;fill:var(--mut)}}
.c-val{{font-size:11px;fill:var(--ink);font-weight:600}}
.tw{{overflow-x:auto;margin:6px 0 16px}}
table{{border-collapse:collapse;width:100%;font-size:13px}}
th,td{{border:1px solid var(--ln);padding:7px 9px;text-align:left;vertical-align:top}}
th{{background:var(--bk);color:#fff;font-weight:700}}
tbody tr:nth-child(even){{background:#FAFAF6}}
.findings{{display:flex;flex-direction:column;gap:9px;margin:8px 0 14px}}
details.f{{border:1px solid var(--ln);border-left:4px solid var(--mut);border-radius:8px;background:#fff}}
details.f.sev-critical{{border-left-color:{BRAND['red']}}}
details.f.sev-high{{border-left-color:{BRAND['orange']}}}
details.f.sev-medium{{border-left-color:{BRAND['blue']}}}
details.f.sev-good{{border-left-color:{BRAND['green']}}}
details.f summary{{cursor:pointer;padding:11px 13px;display:flex;align-items:center;gap:10px;
list-style:none;font-weight:600}}
details.f summary::-webkit-details-marker{{display:none}}
details.f summary::before{{content:"▸";color:var(--mut);transition:.15s}}
details.f[open] summary::before{{transform:rotate(90deg)}}
.chip{{color:#fff;font-size:11px;font-weight:800;padding:2px 8px;border-radius:20px;white-space:nowrap}}
.f-issue{{flex:1}}
.f-body{{padding:2px 14px 14px;border-top:1px solid var(--ln)}}
.f-row{{display:grid;grid-template-columns:92px 1fr;gap:10px;padding:8px 0;border-bottom:1px dashed var(--ln)}}
.f-row:last-child{{border-bottom:none}}
.f-k{{font-weight:800;font-size:11px;text-transform:uppercase;letter-spacing:.4px;color:var(--mut);padding-top:2px}}
.f-v{{white-space:pre-wrap}}
.f-exec{{margin:0;white-space:pre-wrap;background:#0E0E0C;color:#F3F3E9;padding:9px 11px;
border-radius:6px;font:12.5px/1.5 ui-monospace,SFMono-Regular,Menlo,monospace;overflow-x:auto}}
.f-meta{{margin-top:8px;font-size:12px;color:var(--mut)}}
.bar-actions{{margin:2px 0 16px}}
.bar-actions button{{font:inherit;font-size:12px;background:#fff;border:1px solid var(--ln);
border-radius:6px;padding:5px 11px;cursor:pointer}}
.bar-actions button:hover{{border-color:var(--bk)}}
footer{{max-width:1000px;margin:0 auto;padding:20px 22px 40px;color:var(--mut);font-size:12px}}
@media(max-width:620px){{.f-row{{grid-template-columns:1fr}}header.top h1{{font-size:21px}}}}
</style></head><body>
<header class="top"><span class="brand">SEO</span><h1>{header}</h1>{sub}</header>
<nav class="jump">{nav}</nav>
<main>
{sev_chart}
{grid}
<div class="bar-actions"><button data-act="open">Expand all</button>
<button data-act="close">Collapse all</button></div>
{body}
</main>
<footer>Generated by the SEO {esc(title)} skill · Prepared by Bijay · every finding carries Issue · Evidence · Solution · Execution.</footer>
<script>
document.querySelectorAll('.bar-actions button').forEach(function(b){{
  b.addEventListener('click',function(){{
    var open=b.dataset.act==='open';
    document.querySelectorAll('details.f').forEach(function(d){{d.open=open}});
  }});
}});
</script>
</body></html>"""


def _safe_sheet(name, used):
    for ch in '[]:*?/\\':
        name = name.replace(ch, " ")
    name = name.strip()[:31] or "Sheet"
    base, i = name, 2
    while name.lower() in used:
        suffix = f" {i}"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


def render_xlsx(report, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError("openpyxl is required for the XLSX deliverable : "
                          "`pip install openpyxl`, then re-run build_xlsx.py") from e

    sections = report.get("sections") or []
    YL, BK = "FFF5C518", "FF0A0A0A"
    hfill = PatternFill("solid", fgColor=BK)
    hfont = Font(name="Arial", bold=True, color="FFFFFFFF")
    yfont = Font(name="Arial", bold=True, color=BK)
    base = Font(name="Arial")
    wrap = Alignment(wrap_text=True, vertical="top")
    used = set()

    wb = Workbook()
    ov = wb.active
    if ov is None:
        ov = wb.create_sheet("Overview")
    ov.title = _safe_sheet("Overview", used)
    ov["A1"] = "SEO"
    ov["A1"].font = Font(name="Arial", bold=True, size=14, color=BK)
    ov["A1"].fill = PatternFill("solid", fgColor=YL)
    ov["A2"] = f'{report.get("client","Client")} · {report.get("title","Audit")}'
    ov["A2"].font = Font(name="Arial", bold=True, size=12)
    ov["A3"] = f'Period: {report.get("period","")}'
    ov["A3"].font = base
    counts = _sev_counts(sections)
    ov["A5"] = "Findings by severity"
    ov["A5"].font = yfont
    r = 6
    for k in SEV_ORDER:
        if counts[k]:
            ov[f"A{r}"], ov[f"B{r}"] = SEV[k][0], counts[k]
            ov[f"A{r}"].font = base
            ov[f"B{r}"].font = base
            r += 1
    r += 1
    ov[f"A{r}"] = "Section"
    ov[f"B{r}"] = "Findings"
    for c in ("A", "B"):
        ov[f"{c}{r}"].font = hfont
        ov[f"{c}{r}"].fill = hfill
    r += 1
    for s in sections:
        ov[f"A{r}"] = s["title"]
        ov[f"B{r}"] = len(s.get("findings") or [])
        ov[f"A{r}"].font = base
        ov[f"B{r}"].font = base
        r += 1
    ov.column_dimensions["A"].width = 42
    ov.column_dimensions["B"].width = 12

    cols = ["Issue", "Severity", "Evidence", "Solution", "Execution", "Effort", "Priority"]
    widths = [40, 12, 46, 46, 46, 9, 9]
    for s in sections:
        ws = wb.create_sheet(_safe_sheet(s["title"], used))
        for j, c in enumerate(cols, 1):
            cell = ws.cell(1, j, c)
            cell.font = hfont
            cell.fill = hfill
            ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
        ri = 2
        for f in (s.get("findings") or []):
            vals = [f.get("issue"), SEV[_sevkey(f.get("sev"))][0], f.get("evidence"),
                    f.get("solution"), f.get("execution"), f.get("effort", ""), f.get("priority", "")]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(ri, j, v)
                cell.font = base
                cell.alignment = wrap
            ri += 1
        ws.freeze_panes = "A2"

    path = str(path)
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    wb.save(path)
    return path
